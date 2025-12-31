"""
frontend/utils/export_excel.py

Export case results to Excel.
Creates a file with 4 sheets: Summary, Revenue Frame, Efficiency, Configuration.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Dict, Any, Optional, TYPE_CHECKING
from io import BytesIO

if TYPE_CHECKING:
    from pipeline.stages.stage_outputs import (
        ExtractionStageOutput,
        PreDeaStageOutput,
        DeaStageOutput,
        PostDeaStageOutput,
    )

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
DELTA_POS_FILL = PatternFill("solid", fgColor="C6EFCE")
DELTA_NEG_FILL = PatternFill("solid", fgColor="FFC7CE")


class PipelineResultAdapter:
    """
    Adapter för att extrahera data från pipeline-resultat.
    Hanterar både objekt med attribut och dicts.
    """
    def __init__(self, result):
        self.result = result
    
    @property
    def extraction(self):
        return self._get_stage('extraction')
    
    @property
    def pre_dea(self):
        return self._get_stage('pre_dea')
    
    @property
    def dea(self):
        return self._get_stage('dea')
    
    @property
    def post_dea(self):
        return self._get_stage('post_dea')
    
    def _get_stage(self, name: str):
        if hasattr(self.result, name):
            return getattr(self.result, name)
        elif isinstance(self.result, dict):
            return self.result.get(name)
        return None


def create_case_export(
    user_reid: str,
    foretag: str,
    baseline_result,
    case_result,
    ui_config: Dict[str, Any]
) -> BytesIO:
    """
    Creates Excel export of case results.
    
    Args:
        user_reid: User's REId
        foretag: Company name
        baseline_result: Baseline pipeline result (object with .extraction, .pre_dea, etc.)
        case_result: Case pipeline result (object with .extraction, .pre_dea, etc.)
        ui_config: UI configuration (parameters)
    
    Returns:
        BytesIO with Excel file ready for download
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Wrap results for unified access
    baseline = PipelineResultAdapter(baseline_result)
    case = PipelineResultAdapter(case_result)
    
    _create_summary_sheet(wb, user_reid, foretag, baseline, case)
    _create_intaktsram_sheet(wb, baseline, case)
    _create_efficiency_sheet(wb, baseline, case)
    _create_config_sheet(wb, ui_config, case)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def get_export_filename(user_reid: str) -> str:
    """Generates filename with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"regumetrica_case_{user_reid}_{timestamp}.xlsx"


def _create_summary_sheet(
    wb: Workbook,
    user_reid: str,
    foretag: str,
    baseline: PipelineResultAdapter,
    case: PipelineResultAdapter
):
    """Creates summary sheet with baseline vs case comparison."""
    ws = wb.create_sheet("Summary")
    
    # Header
    ws['A1'] = "Regumetrica - Case Results"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = "Company:"
    ws['B3'] = f"{foretag} ({user_reid})"
    ws['A4'] = "Export date:"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Metadata
    ws['A6'] = "Case metadata"
    ws['A6'].font = Font(bold=True)
    
    case_pre_dea = case.pre_dea
    case_dea = case.dea
    
    ws['A7'] = "CAPEX method:"
    ws['B7'] = _get_attr(case_pre_dea, 'capex_method', 'baseline')
    ws['A8'] = "DEA method:"
    ws['B8'] = _get_attr(case_dea, 'dea_method', 'baseline')
    ws['A9'] = "DEA executed:"
    ws['B9'] = "Yes" if _get_attr(case_dea, 'dea_executed', False) else "No"
    
    # Comparison table
    ws['A11'] = "Comparison: Baseline vs Case"
    ws['A11'].font = Font(bold=True)
    
    headers = ['Variable', 'Baseline', 'Case', 'Delta', 'Delta %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=12, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # Get revenue frame data
    baseline_ir = _get_attr(baseline.post_dea, 'user_intaktsram', None)
    case_ir = _get_attr(case.post_dea, 'user_intaktsram', None)
    baseline_effkrav = _get_attr(baseline.post_dea, 'user_effkrav_proc', 0)
    case_effkrav = _get_attr(case.post_dea, 'user_effkrav_proc', 0)
    
    # Data rows
    comparison_data = [
        ('Kapitalkostnad_Total', 'tkr'),
        ('Paverkbara_Periodsumma', 'tkr'),
        ('Opaverkbara_Kostnader', 'tkr'),
        ('Intaktsram_Total', 'tkr'),
    ]
    
    row = 13
    for key, unit in comparison_data:
        baseline_val = _get_series_value(baseline_ir, key)
        case_val = _get_series_value(case_ir, key)
        delta, delta_pct = _calc_delta(baseline_val, case_val)
        
        ws.cell(row=row, column=1, value=f"{key} ({unit})")
        ws.cell(row=row, column=2, value=baseline_val).number_format = '#,##0'
        ws.cell(row=row, column=3, value=case_val).number_format = '#,##0'
        
        delta_cell = ws.cell(row=row, column=4, value=delta)
        delta_cell.number_format = '#,##0'
        if delta and delta > 0:
            delta_cell.fill = DELTA_POS_FILL
        elif delta and delta < 0:
            delta_cell.fill = DELTA_NEG_FILL
        
        pct_cell = ws.cell(row=row, column=5, value=delta_pct)
        pct_cell.number_format = '0.00%'
        
        row += 1
    
    # Efficiency requirement
    ws.cell(row=row, column=1, value="Effkrav_proc")
    ws.cell(row=row, column=2, value=baseline_effkrav).number_format = '0.00%'
    ws.cell(row=row, column=3, value=case_effkrav).number_format = '0.00%'
    delta_effkrav = case_effkrav - baseline_effkrav if baseline_effkrav and case_effkrav else None
    ws.cell(row=row, column=4, value=delta_effkrav).number_format = '0.00%'
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12


def _create_intaktsram_sheet(
    wb: Workbook, 
    baseline: PipelineResultAdapter, 
    case: PipelineResultAdapter
):
    """Creates revenue frame sheet with all companies."""
    ws = wb.create_sheet("Revenue Frame")
    
    baseline_df = _get_attr(baseline.post_dea, 'all_intaktsram', None)
    case_df = _get_attr(case.post_dea, 'all_intaktsram', None)
    
    if baseline_df is None or case_df is None:
        ws['A1'] = "Data missing"
        return
    
    if not isinstance(baseline_df, pd.DataFrame) or not isinstance(case_df, pd.DataFrame):
        ws['A1'] = "Data is not DataFrame"
        return
    
    baseline_df = baseline_df.copy()
    case_df = case_df.copy()
    
    # Filter to REL companies
    if 'REId' in baseline_df.columns:
        baseline_df = baseline_df[baseline_df['REId'].str.startswith('REL', na=False)].copy()
    if 'REId' in case_df.columns:
        case_df = case_df[case_df['REId'].str.startswith('REL', na=False)].copy()
    
    # Select columns
    keep_cols = ['REId', 'Kapitalkostnad_Total', 'Paverkbara_Periodsumma', 
                 'Opaverkbara_Kostnader', 'Intaktsram_Total']
    
    baseline_cols = [c for c in keep_cols if c in baseline_df.columns]
    case_cols = [c for c in keep_cols if c in case_df.columns]
    
    if not baseline_cols or not case_cols:
        ws['A1'] = "Columns missing"
        return
    
    baseline_df = baseline_df[baseline_cols].copy()
    case_df = case_df[case_cols].copy()
    
    # Merge and calculate delta
    merged = baseline_df.merge(case_df, on='REId', suffixes=('_baseline', '_case'))
    
    for col in ['Kapitalkostnad_Total', 'Paverkbara_Periodsumma', 'Opaverkbara_Kostnader', 'Intaktsram_Total']:
        if f'{col}_baseline' in merged.columns and f'{col}_case' in merged.columns:
            merged[f'{col}_Delta'] = merged[f'{col}_case'] - merged[f'{col}_baseline']
            merged[f'{col}_Delta%'] = merged[f'{col}_Delta'] / merged[f'{col}_baseline'].replace(0, float('nan'))
    
    _write_dataframe_to_sheet(ws, merged)


def _create_efficiency_sheet(
    wb: Workbook, 
    baseline: PipelineResultAdapter, 
    case: PipelineResultAdapter
):
    """Creates efficiency sheet."""
    ws = wb.create_sheet("Efficiency")
    
    baseline_dea = baseline.dea
    case_dea = case.dea
    
    ws['A1'] = "Efficiency Analysis"
    ws['A1'].font = Font(bold=True, size=12)
    
    # DEA metadata
    ws['A3'] = "DEA Configuration"
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = "Method:"
    ws['B4'] = _get_attr(case_dea, 'dea_method', 'baseline')
    ws['A5'] = "Executed:"
    ws['B5'] = "Yes" if _get_attr(case_dea, 'dea_executed', False) else "No"
    
    # Efficiency scores if available
    case_scores = _get_attr(case_dea, 'efficiency_scores', None)
    if case_scores is not None and isinstance(case_scores, pd.DataFrame):
        ws['A7'] = "Efficiency Scores"
        ws['A7'].font = Font(bold=True)
        _write_dataframe_to_sheet(ws, case_scores, start_row=8)


def _create_config_sheet(
    wb: Workbook,
    ui_config: Dict[str, Any],
    case: PipelineResultAdapter
):
    """Creates configuration sheet with all parameters."""
    ws = wb.create_sheet("Configuration")
    
    ws['A1'] = "Case Configuration"
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for module_name, module_config in ui_config.items():
        ws.cell(row=row, column=1, value=module_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        if isinstance(module_config, dict):
            for key, value in module_config.items():
                # Skip binary data
                if isinstance(value, bytes):
                    ws.cell(row=row, column=2, value=key)
                    ws.cell(row=row, column=3, value="[binary data]")
                elif isinstance(value, (dict, list)):
                    ws.cell(row=row, column=2, value=key)
                    ws.cell(row=row, column=3, value=str(value)[:100])
                else:
                    ws.cell(row=row, column=2, value=key)
                    ws.cell(row=row, column=3, value=value)
                row += 1
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50


# --- Helper functions ---

def _get_attr(obj, attr: str, default=None):
    """Safely get attribute from object or dict."""
    if obj is None:
        return default
    if hasattr(obj, attr):
        return getattr(obj, attr)
    elif isinstance(obj, dict):
        return obj.get(attr, default)
    return default


def _get_series_value(series, key, default=None):
    """Safely get value from pandas Series."""
    if series is None:
        return default
    if isinstance(series, pd.Series) and key in series.index:
        return series[key]
    return default


def _calc_delta(baseline_val, case_val):
    """Calculate delta and percentage change."""
    if baseline_val is None or case_val is None:
        return None, None
    delta = case_val - baseline_val
    delta_pct = delta / baseline_val if baseline_val != 0 else None
    return delta, delta_pct


def _write_dataframe_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    """Write DataFrame to worksheet with headers."""
    # Headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    # Data
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, float):
                cell.number_format = '#,##0.00'