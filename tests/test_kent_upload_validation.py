"""
Tests for KENT upload validation (diagnose_kent_upload) and category matching.

Builds minimal KENT-shaped .xlsx files in tmp_path (independent of the real
template) and asserts the structured deviation report: blocking errors for
non-template files, and proceed-but-flag warnings (with assumptions) for
deviations like unknown categories or missing years.
"""
import openpyxl
import pytest

from calculations.capex.kent_capbase_prep import (
    diagnose_kent_upload,
    category_match_kind,
)

NID = 999  # arbitrary; network_id only labels id_network, no baseline dependency

# Normvärde header on Excel row 1 (no title row); Övriga/Investeringar carry a
# title row above the header on row 2 — exactly like the real template.
_NORM_HEADERS = ["Anl.-kategori", "Kod", "Typ av anläggning", "Antal",
                 "Rådighet", "Ursprungligen tagen i bruk", "NUAV"]
_OVRIGA_HEADERS = ["Ansk", "Bokf", "Annat", "Anl.kategori", "Typ av anläggning",
                   "Antal", "Ursprungligen tagen i bruk", "Rådighet", "NUAV 2022"]
_INVEST_HEADERS = ["Investering / Utrangering", "Halvår", "Anl.kategori",
                   "Typ av anläggning", "Antal", "Ursprungligen tagen i bruk",
                   "Totalt i kronor"]


def _write_kent(path, norm_rows):
    """norm_rows: list of (kat, kod, typ, antal, radighet, year, nuav)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normvärde"
    for c, h in enumerate(_NORM_HEADERS, start=1):
        ws.cell(1, c, h)
    for r, row in enumerate(norm_rows, start=2):
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(r, c, val)

    ws2 = wb.create_sheet("Övriga värderingsmetoder")
    ws2.cell(1, 1, "Övriga värderingsmetoder")  # title row
    for c, h in enumerate(_OVRIGA_HEADERS, start=1):
        ws2.cell(2, c, h)

    ws3 = wb.create_sheet("Investeringar_Utrangeringar")
    ws3.cell(1, 1, "Planerade investeringar/utrangeringar")  # title row
    for c, h in enumerate(_INVEST_HEADERS, start=1):
        ws3.cell(2, c, h)

    wb.save(path)
    return str(path)


# A clean, ordinarie pair of official categories.
_CLEAN_ROWS = [
    ("Nätstation", 13, "Standard", 5, "Ägd", 2010, 10_000_000),
    ("Mätare", 12, "Elmätare", 5000, "Ägd", 2018, 20_000_000),
]


def test_clean_file_ok_no_warnings(tmp_path):
    p = _write_kent(tmp_path / "clean.xlsx", _CLEAN_ROWS)
    rep = diagnose_kent_upload(p, NID)
    assert rep["ok"] is True
    assert rep["errors"] == []
    assert rep["warnings"] == []
    assert rep["summary"]["n_components"] == 2


def test_unrelated_file_is_blocking_error(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active.cell(1, 1, "foo")
    p = tmp_path / "unrelated.xlsx"
    wb.save(p)
    rep = diagnose_kent_upload(str(p), NID)
    assert rep["ok"] is False
    assert any("template" in e["issue"].lower() for e in rep["errors"])


def test_corrupt_non_xlsx_is_blocking_error(tmp_path):
    p = tmp_path / "corrupt.xlsx"
    p.write_bytes(b"this is definitely not an excel file")
    rep = diagnose_kent_upload(str(p), NID)
    assert rep["ok"] is False
    assert any("Excel" in e["issue"] for e in rep["errors"])


def test_unknown_category_warns_with_transformator_assumption(tmp_path):
    rows = _CLEAN_ROWS + [("Helt Okänd Grej", 1, "X", 1, "Ägd", 2015, 1_000_000)]
    p = _write_kent(tmp_path / "unknown_cat.xlsx", rows)
    rep = diagnose_kent_upload(p, NID)
    assert rep["ok"] is True  # deviation, not blocking
    w = [x for x in rep["warnings"] if "category" in x["issue"].lower()]
    assert w, "expected an unknown-category warning"
    assert "Transformator" in w[0]["assumption"]


def test_missing_year_warns(tmp_path):
    rows = _CLEAN_ROWS + [("Nätstation", 13, "Standard", 1, "Ägd", None, 1_000_000)]
    p = _write_kent(tmp_path / "no_year.xlsx", rows)
    rep = diagnose_kent_upload(p, NID)
    assert rep["ok"] is True
    assert any("year" in x["issue"].lower() for x in rep["warnings"])


def test_zero_nuav_warns(tmp_path):
    rows = _CLEAN_ROWS + [("Mätare", 12, "Elmätare", 1, "Ägd", 2015, 0)]
    p = _write_kent(tmp_path / "zero_nuav.xlsx", rows)
    rep = diagnose_kent_upload(p, NID)
    assert any("NUAV" in x["issue"] for x in rep["warnings"])


@pytest.mark.parametrize("text,kind", [
    ("Transformator", "exact"),
    ("transformator", "exact"),
    ("Ledning med en spänning om 220 kV eller mer, med undantag för luftledning, linjekoncession", "exact"),
    ("Kabel", "substring"),
    ("Jordkabel City", "substring"),
    ("Helt Okänd Grej", "default"),
    ("", "empty"),
    (None, "empty"),
])
def test_category_match_kind(text, kind):
    assert category_match_kind(text) == kind
