"""
pipeline/export_excel.py

Professional Excel export from PipelineResult.

Structure:
  1. Summary          -- Polished report with sections per module (case vs baseline)
  2. Capital Base Detail -- Half-year NUAV/Dep/Return per category for user's company
  3. All Companies - Revenue Frame  -- Raw 148-row data
  4. All Companies - DEA            -- Raw 148-row data
  5. All Companies - Costs          -- Raw 148-row data
  6. All Companies - Incentives     -- Raw 148-row data
  7. Configuration                  -- Audit trail of all parameters
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple, TYPE_CHECKING
from io import BytesIO

if TYPE_CHECKING:
    from pipeline.core import PipelineResult

from config.column_names import (
    COL_REID, COL_COMPANY_NAME, COL_ID_NETWORK,
    COL_CAPITAL_COST_2024, COL_CAPITAL_COST_2025,
    COL_CAPITAL_COST_2026, COL_CAPITAL_COST_2027, COL_CAPITAL_COST_PERIOD,
    COL_DEPRECIATION_2024, COL_DEPRECIATION_2025,
    COL_DEPRECIATION_2026, COL_DEPRECIATION_2027, COL_DEPRECIATION_PERIOD,
    COL_RETURN_2024, COL_RETURN_2025,
    COL_RETURN_2026, COL_RETURN_2027, COL_RETURN_PERIOD,
    COL_CONTROLLABLE_AVG, COL_CONTROLLABLE_PERIOD, COL_CONTROLLABLE_BEFORE,
    COL_CONTROLLABLE_IN_RF,
    COL_OPEX_BEFORE, COL_OPEX_AFTER, COL_OPEX_EFF_DEDUCTION, COL_OPEX_SHARE,
    COL_CAPEX_BEFORE, COL_CAPEX_AFTER, COL_CAPEX_EFF_DEDUCTION, COL_CAPEX_SHARE,
    COL_EFFICIENCY_DEDUCTION,
    COL_NON_CONTROLLABLE, COL_FLEXIBILITY, COL_INTERRUPTION, COL_STATE_DEDUCTION,
    COL_REVENUE_FRAME, COL_METHOD_USED,
    COL_DEA_EFFICIENCY, COL_DEA_SUPER_EFF, COL_DEA_POTENTIAL, COL_IS_OUTLIER,
    COL_EFF_REQ_ANNUAL,
    COL_QUALITY_INCENTIVE, COL_NETLOSS_INCENTIVE, COL_LOAD_INCENTIVE,
    COL_INCENTIVE_TOTAL, COL_MISSING_INCENTIVE,
    COL_TOTEX, COL_CU, COL_MW, COL_NS, COL_MWH_LOW, COL_MWH_HIGH,
)
from config.asset_categories import CATEGORY_BY_CODE, get_category_short_name
from config.time_codes import (
    TIMECODE_TO_HALFYEAR,
    PERIOD_2024_2027_CODES,
)


# =============================================================================
# EXCEL STYLE CONSTANTS
# =============================================================================

# Colors (hex without #, for openpyxl)
_CLR_PRIMARY = "2563EB"
_CLR_PRIMARY_LIGHT = "DBEAFE"
_CLR_BG_SUBTLE = "F1F5F9"
_CLR_BG_MUTED = "E2E8F0"
_CLR_TEXT_PRIMARY = "0F172A"
_CLR_TEXT_SECONDARY = "475569"
_CLR_SUCCESS = "059669"
_CLR_ERROR = "DC2626"
_CLR_WHITE = "FFFFFF"

# Reusable style objects
_FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=_CLR_PRIMARY)
_FONT_COMPANY = Font(name="Calibri", size=12, bold=True, color=_CLR_TEXT_PRIMARY)
_FONT_META = Font(name="Calibri", size=10, color=_CLR_TEXT_SECONDARY)
_FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=_CLR_TEXT_PRIMARY)
_FONT_COL_HEADER = Font(name="Calibri", size=10, bold=True, color=_CLR_WHITE)
_FONT_DATA = Font(name="Calibri", size=10, color=_CLR_TEXT_PRIMARY)
_FONT_DATA_BOLD = Font(name="Calibri", size=10, bold=True, color=_CLR_TEXT_PRIMARY)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color=_CLR_WHITE)
_FONT_DELTA_POS = Font(name="Calibri", size=10, color=_CLR_SUCCESS)
_FONT_DELTA_NEG = Font(name="Calibri", size=10, color=_CLR_ERROR)

_FILL_PRIMARY = PatternFill("solid", fgColor=_CLR_PRIMARY)
_FILL_SECTION = PatternFill("solid", fgColor=_CLR_BG_SUBTLE)
_FILL_USER_ROW = PatternFill("solid", fgColor=_CLR_PRIMARY_LIGHT)

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_CLR_BG_MUTED),
)
_BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color=_CLR_TEXT_PRIMARY),
)

# Number format strings (Excel locale-aware)
_FMT_TKR = '#,##0'
_FMT_TKR_SIGN = '+#,##0;-#,##0;"-"'
_FMT_PCT = '0.00%'
_FMT_PCT_SIGN = '+0.00%;-0.00%'
_FMT_RATIO = '0.0000'
_FMT_INT = '0'

# Column widths by type
_W_LABEL = 35
_W_TKR = 16
_W_PCT = 13
_W_REID = 12
_W_NAME = 32
_W_SMALL = 10
_W_BOOL = 10


# =============================================================================
# UTILITY: HEADER BLOCK
# =============================================================================

def _write_header_block(
    ws,
    company_name: str,
    user_reid: str,
    case_name: str,
    start_row: int = 1,
) -> int:
    """Write Regumetrica header block. Returns next available row."""
    r = start_row
    cell = ws.cell(row=r, column=1, value="Regumetrica")
    cell.font = _FONT_TITLE
    r += 1

    cell = ws.cell(row=r, column=1, value=f"{company_name} ({user_reid})")
    cell.font = _FONT_COMPANY
    r += 1

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta = f"Case: {case_name}" if case_name else "Baseline"
    cell = ws.cell(row=r, column=1, value=f"{meta}  |  Exported: {ts}")
    cell.font = _FONT_META
    r += 2  # blank row after header

    return r


# =============================================================================
# UTILITY: SECTION HEADER
# =============================================================================

def _write_section_header(ws, row: int, title: str, n_cols: int = 5) -> int:
    """Write a section header row with subtle fill. Returns next row."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _FILL_SECTION
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _FONT_SECTION
    cell.fill = _FILL_SECTION
    return row + 1


# =============================================================================
# UTILITY: COLUMN HEADERS
# =============================================================================

def _write_col_headers(ws, row: int, headers: List[str]) -> int:
    """Write styled column header row. Returns next row."""
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.font = _FONT_COL_HEADER
        cell.fill = _FILL_PRIMARY
        cell.alignment = _ALIGN_CENTER
    return row + 1


# =============================================================================
# UTILITY: WRITE DATA ROW
# =============================================================================

def _write_data_row(
    ws,
    row: int,
    values: list,
    formats: Optional[list] = None,
    is_total: bool = False,
    delta_cols: Optional[List[int]] = None,
) -> int:
    """Write one data row. Returns next row.

    Args:
        formats: list of format strings per column (None = no format).
        is_total: if True, apply total row styling (bold primary fill).
        delta_cols: 0-based column indices where positive=green, negative=red font.
    """
    delta_cols = delta_cols or []

    for c, val in enumerate(values):
        cell = ws.cell(row=row, column=c + 1)

        if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
            cell.value = None
        elif isinstance(val, (int, float, np.integer, np.floating)):
            cell.value = float(val)
        else:
            cell.value = str(val)

        # Alignment
        if isinstance(val, str) or c == 0:
            cell.alignment = _ALIGN_LEFT
        else:
            cell.alignment = _ALIGN_RIGHT

        # Number format
        if formats and c < len(formats) and formats[c]:
            cell.number_format = formats[c]

        # Total row styling
        if is_total:
            cell.font = _FONT_TOTAL
            cell.fill = _FILL_PRIMARY
            cell.border = _BOTTOM_BORDER
        else:
            cell.font = _FONT_DATA
            cell.border = _THIN_BORDER

            # Delta coloring
            if c in delta_cols and isinstance(val, (int, float, np.integer, np.floating)):
                if val > 0.5:
                    cell.font = _FONT_DELTA_POS
                elif val < -0.5:
                    cell.font = _FONT_DELTA_NEG

    return row + 1


# =============================================================================
# UTILITY: WRITE RAW DATAFRAME
# =============================================================================

def _write_raw_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    highlight_reid: Optional[str] = None,
    column_formats: Optional[Dict[str, str]] = None,
) -> int:
    """Write a flat DataFrame with styled headers. Returns next row after data."""
    column_formats = column_formats or {}

    # Headers
    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=str(col_name))
        cell.font = _FONT_COL_HEADER
        cell.fill = _FILL_PRIMARY
        cell.alignment = _ALIGN_CENTER

    # Data
    reid_col_idx = None
    if COL_REID in df.columns:
        reid_col_idx = list(df.columns).index(COL_REID)

    for r_idx, row_data in enumerate(df.itertuples(index=False), start_row + 1):
        is_user_row = False
        if highlight_reid and reid_col_idx is not None:
            reid_val = row_data[reid_col_idx]
            is_user_row = (str(reid_val) == str(highlight_reid))

        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=r_idx, column=c_idx + 1)
            col_name = df.columns[c_idx]

            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = None
            elif isinstance(value, (bool, np.bool_)):
                cell.value = "Yes" if value else "No"
            elif isinstance(value, (int, float, np.integer, np.floating)):
                cell.value = float(value)
            else:
                cell.value = str(value)

            # Number format
            fmt = column_formats.get(col_name)
            if fmt:
                cell.number_format = fmt

            # Alignment
            if isinstance(value, (int, float, np.integer, np.floating)):
                cell.alignment = _ALIGN_RIGHT
            else:
                cell.alignment = _ALIGN_LEFT

            cell.font = _FONT_DATA

            # Highlight user row
            if is_user_row:
                cell.fill = _FILL_USER_ROW

    return start_row + 1 + len(df)


def _set_column_widths(ws, widths: List[Tuple[int, float]]):
    """Set column widths. widths = [(col_1based, width), ...]."""
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


# =============================================================================
# UTILITY: SAFE VALUE EXTRACTION
# =============================================================================

def _safe(series_or_dict, key, default=0.0):
    """Safely extract a value from a Series or dict."""
    if series_or_dict is None:
        return default
    val = series_or_dict.get(key, default) if hasattr(series_or_dict, 'get') else default
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    return float(val)


def _safe_str(series_or_dict, key, default=""):
    """Safely extract a string value."""
    if series_or_dict is None:
        return default
    val = series_or_dict.get(key, default) if hasattr(series_or_dict, 'get') else default
    if val is None:
        return default
    return str(val)


def _delta(case_val, bl_val):
    """Compute absolute delta."""
    return case_val - bl_val


def _delta_pct(case_val, bl_val):
    """Compute percentage delta (as decimal for Excel %, e.g. 0.05 = 5%)."""
    if abs(bl_val) < 0.01:
        return None
    return (case_val - bl_val) / bl_val


# =============================================================================
# DATA HELPERS
# =============================================================================

def _get_company_names(result) -> Dict[str, str]:
    """Build REId -> company_name lookup from pre_dea data."""
    df = result.pre_dea.df_all_companies
    if COL_COMPANY_NAME in df.columns and COL_REID in df.columns:
        return dict(zip(df[COL_REID], df[COL_COMPANY_NAME]))
    return {}


def _compute_rank(dea_results: pd.DataFrame, user_reid: str) -> Tuple[Optional[int], int]:
    """User's rank by efficiency (1 = best). Returns (rank, total)."""
    if dea_results is None or dea_results.empty:
        return None, 0
    df = dea_results[[COL_REID, COL_DEA_EFFICIENCY]].dropna(subset=[COL_DEA_EFFICIENCY])
    df = df.sort_values(COL_DEA_EFFICIENCY, ascending=False).reset_index(drop=True)
    total = len(df)
    matches = df.index[df[COL_REID] == user_reid]
    if len(matches) == 0:
        return None, total
    return int(matches[0]) + 1, total


def _get_depreciation_return(case_result, baseline_result, user_reid: str) -> Dict[str, Dict[str, float]]:
    """Extract depreciation and return period totals (same logic as diagram_data).

    Returns dict with keys 'depreciation' and 'return', each having 'case' and 'baseline'.
    Also returns per-year return if available from pre_dea.
    """
    from calculations.capex.wacc_calculations import BASELINE_WACC

    capex_method = case_result.pre_dea.capex_method
    sdf_ir = baseline_result.baseline.sdf_ir

    # Baseline: always from SDF
    bl_dep, bl_ret = 0.0, 0.0
    if sdf_ir is not None and not sdf_ir.empty:
        user_mask = sdf_ir['REId'] == user_reid
        if user_mask.any():
            user_row = sdf_ir[user_mask].iloc[0]
            if COL_RETURN_PERIOD in sdf_ir.columns:
                bl_ret = float(pd.to_numeric(user_row.get(COL_RETURN_PERIOD, 0), errors='coerce') or 0)
            if COL_CAPITAL_COST_PERIOD in sdf_ir.columns:
                bl_cap = float(pd.to_numeric(user_row.get(COL_CAPITAL_COST_PERIOD, 0), errors='coerce') or 0)
                bl_dep = bl_cap - bl_ret

    # Case: from SDF (baseline mode) or pre_dea (parameter_change mode)
    if capex_method == 'baseline':
        c_dep, c_ret = bl_dep, bl_ret
    else:
        df_all = case_result.pre_dea.df_all_companies
        user_mask = df_all[COL_REID] == user_reid
        if user_mask.any():
            user_row = df_all[user_mask].iloc[0]
            c_dep = 0.0
            c_ret = 0.0
            if COL_DEPRECIATION_PERIOD in df_all.columns:
                c_dep = float(user_row.get(COL_DEPRECIATION_PERIOD, 0) or 0)
            elif all(f'depreciation_{y}' in df_all.columns for y in [2024, 2025, 2026, 2027]):
                c_dep = sum(float(user_row.get(f'depreciation_{y}', 0) or 0) for y in [2024, 2025, 2026, 2027])
            if COL_RETURN_PERIOD in df_all.columns:
                c_ret = float(user_row.get(COL_RETURN_PERIOD, 0) or 0)
            elif all(f'return_on_assets_{y}' in df_all.columns for y in [2024, 2025, 2026, 2027]):
                c_ret = sum(float(user_row.get(f'return_on_assets_{y}', 0) or 0) for y in [2024, 2025, 2026, 2027])
        else:
            c_dep, c_ret = bl_dep, bl_ret

    return {
        'depreciation': {'case': c_dep, 'baseline': bl_dep},
        'return': {'case': c_ret, 'baseline': bl_ret},
    }


def _get_return_per_year(case_result, baseline_result, user_reid: str) -> Dict[str, Dict[str, float]]:
    """Extract return on assets per year from pre_dea/SDF.

    Returns {year_str: {'case': val, 'baseline': val}} for 2024-2027 + 'Period'.
    """
    years = [2024, 2025, 2026, 2027]
    result = {}

    # Baseline per year from SDF
    sdf_ir = baseline_result.baseline.sdf_ir
    bl_yearly = {}
    if sdf_ir is not None and not sdf_ir.empty:
        user_mask = sdf_ir['REId'] == user_reid
        if user_mask.any():
            user_row = sdf_ir[user_mask].iloc[0]
            for y in years:
                col = f'return_on_assets_{y}'
                if col in sdf_ir.columns:
                    bl_yearly[y] = float(pd.to_numeric(user_row.get(col, 0), errors='coerce') or 0)
                else:
                    bl_yearly[y] = 0.0
        else:
            bl_yearly = {y: 0.0 for y in years}
    else:
        bl_yearly = {y: 0.0 for y in years}

    # Case per year from pre_dea (or same as baseline if capex_method == 'baseline')
    capex_method = case_result.pre_dea.capex_method
    if capex_method == 'baseline':
        c_yearly = dict(bl_yearly)
    else:
        df_all = case_result.pre_dea.df_all_companies
        c_yearly = {}
        user_mask = df_all[COL_REID] == user_reid
        if user_mask.any():
            user_row = df_all[user_mask].iloc[0]
            for y in years:
                col = f'return_on_assets_{y}'
                if col in df_all.columns:
                    c_yearly[y] = float(user_row.get(col, 0) or 0)
                else:
                    c_yearly[y] = 0.0
        else:
            c_yearly = {y: 0.0 for y in years}

    for y in years:
        result[str(y)] = {'case': c_yearly.get(y, 0.0), 'baseline': bl_yearly.get(y, 0.0)}

    result['Period'] = {
        'case': sum(c_yearly.values()),
        'baseline': sum(bl_yearly.values()),
    }
    return result


def _load_category_data(case_result, baseline_result, user_id_network):
    """Load category-level data for user's company (case + baseline).

    Returns (case_cat_df, baseline_cat_df) -- either may be None.
    """
    from pipeline.result_helpers import (
        load_baseline_category_data,
        get_case_category_data,
    )
    baseline_cat = load_baseline_category_data(user_id_network)
    case_cat = get_case_category_data(case_result, user_id_network)
    if case_cat is None or case_cat.empty:
        case_cat = baseline_cat.copy() if baseline_cat is not None else None
    return case_cat, baseline_cat


def _aggregate_period_for_component(df, col_ord, col_tail):
    """Aggregate half-year data to period totals per category for one component."""
    if df is None or df.empty:
        return {}
    for col in [col_ord, col_tail]:
        if col not in df.columns:
            df[col] = 0.0
    agg = df.groupby('cat_encode').agg({col_ord: 'sum', col_tail: 'sum'}).reset_index()
    result = {}
    for _, row in agg.iterrows():
        ce = int(row['cat_encode'])
        o = float(row[col_ord])
        t = float(row[col_tail])
        result[ce] = (o, t, o + t)
    return result


# =============================================================================
# SHEET 1: SUMMARY
# =============================================================================

def _build_summary_sheet(
    wb: Workbook,
    case_result,
    baseline_result,
    ui_config: Dict[str, Any],
    company_name: str,
    user_reid: str,
    case_name: str,
):
    ws = wb.create_sheet("Summary")
    n_cols = 5  # Component | Case | Baseline | Delta | Delta%

    row = _write_header_block(ws, company_name, user_reid, case_name)

    case_ir = case_result.post_dea.user_revenue_frame
    bl_ir = baseline_result.post_dea.user_revenue_frame

    # Extract depreciation/return from SDF/pre_dea (not in user_revenue_frame)
    dep_ret = _get_depreciation_return(case_result, baseline_result, user_reid)

    # ---- Section A: Revenue Frame ----
    row = _write_section_header(ws, row, "Revenue Frame (Period 2024-2027)", n_cols)
    headers = ["Component", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"]
    row = _write_col_headers(ws, row, headers)

    rf_fmts = [None, _FMT_TKR, _FMT_TKR, _FMT_TKR_SIGN, _FMT_PCT_SIGN]

    # Components from revenue frame Series, ordered to match waterfall
    rf_components_before_dep = [
        ("40.1.1 Controllable OPEX", COL_CONTROLLABLE_IN_RF, COL_CONTROLLABLE_PERIOD),
        ("40.2.1 Non-controllable OPEX", COL_NON_CONTROLLABLE, None),
    ]
    rf_components_after_dep = [
        ("50.4.1 Efficiency OPEX adj.", COL_OPEX_EFF_DEDUCTION, None),
        ("50.4.2 Efficiency CAPEX adj.", COL_CAPEX_EFF_DEDUCTION, None),
        ("30.5.2 Incentive adjustment", COL_INCENTIVE_TOTAL, None),
        ("40.1.2 Flexibility services", COL_FLEXIBILITY, None),
        ("Interruption compensation", COL_INTERRUPTION, None),
        ("State aid deduction", COL_STATE_DEDUCTION, None),
    ]

    # Controllable + Non-controllable first
    for label, col_key, fallback_key in rf_components_before_dep:
        c_val = _safe(case_ir, col_key, _safe(case_ir, fallback_key) if fallback_key else 0.0)
        b_val = _safe(bl_ir, col_key, _safe(bl_ir, fallback_key) if fallback_key else 0.0)
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        row = _write_data_row(ws, row, [label, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3])

    # Depreciation and Return (from SDF/pre_dea, not revenue frame)
    for label, dep_ret_key in [("20.1 Depreciation", "depreciation"), ("30.1 Return (WACC)", "return")]:
        c_val = dep_ret[dep_ret_key]['case']
        b_val = dep_ret[dep_ret_key]['baseline']
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        row = _write_data_row(ws, row, [label, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3])

    # Efficiency adjustments, incentives, other
    for label, col_key, fallback_key in rf_components_after_dep:
        c_val = _safe(case_ir, col_key, _safe(case_ir, fallback_key) if fallback_key else 0.0)
        b_val = _safe(bl_ir, col_key, _safe(bl_ir, fallback_key) if fallback_key else 0.0)
        # Negate deductions for display (they reduce the frame)
        if "deduction" in col_key or col_key == COL_STATE_DEDUCTION:
            c_val = -c_val
            b_val = -b_val
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        row = _write_data_row(ws, row, [label, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3])

    # Total row
    c_total = _safe(case_ir, COL_REVENUE_FRAME)
    b_total = _safe(bl_ir, COL_REVENUE_FRAME)
    d_total = _delta(c_total, b_total)
    dp_total = _delta_pct(c_total, b_total)
    row = _write_data_row(ws, row, ["TOTAL REVENUE FRAME", c_total, b_total, d_total, dp_total],
                          formats=rf_fmts, is_total=True)
    row += 1  # blank

    # ---- Section B: M1 Asset Base (NUAV by category) ----
    user_id_network = getattr(case_result.pre_dea, 'user_id_network', None)
    if user_id_network is not None:
        case_cat, bl_cat = _load_category_data(case_result, baseline_result, user_id_network)
        if case_cat is not None and bl_cat is not None:
            case_nuav = _aggregate_period_for_component(case_cat, 'nuav_ord', 'nuav_tail')
            bl_nuav = _aggregate_period_for_component(bl_cat, 'nuav_ord', 'nuav_tail')

            row = _write_section_header(ws, row, "M1 Asset Base Valuation (NUAV, period total)", n_cols)
            row = _write_col_headers(ws, row, ["Category", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"])

            active_cats = sorted(set(case_nuav.keys()) | set(bl_nuav.keys()))
            total_case, total_bl = 0.0, 0.0

            for ce in active_cats:
                c_o, c_t, c_tot = case_nuav.get(ce, (0, 0, 0))
                b_o, b_t, b_tot = bl_nuav.get(ce, (0, 0, 0))
                if abs(c_tot) < 0.01 and abs(b_tot) < 0.01:
                    continue
                total_case += c_tot
                total_bl += b_tot
                name = get_category_short_name(ce)
                d = _delta(c_tot, b_tot)
                dp = _delta_pct(c_tot, b_tot)
                row = _write_data_row(ws, row, [name, c_tot, b_tot, d, dp],
                                      formats=rf_fmts, delta_cols=[3])

            d = _delta(total_case, total_bl)
            dp = _delta_pct(total_case, total_bl)
            row = _write_data_row(ws, row, ["TOTAL NUAV", total_case, total_bl, d, dp],
                                  formats=rf_fmts, is_total=True)
            row += 1

            # ---- Section C: M2 Depreciation ----
            case_dep = _aggregate_period_for_component(case_cat, 'dep_ord', 'dep_tail')
            bl_dep = _aggregate_period_for_component(bl_cat, 'dep_ord', 'dep_tail')

            if case_dep or bl_dep:
                row = _write_section_header(ws, row, "M2 Depreciation (period total)", n_cols)
                row = _write_col_headers(ws, row, ["Category", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"])

                total_case_dep, total_bl_dep = 0.0, 0.0
                for ce in active_cats:
                    c_o, c_t, c_tot = case_dep.get(ce, (0, 0, 0))
                    b_o, b_t, b_tot = bl_dep.get(ce, (0, 0, 0))
                    if abs(c_tot) < 0.01 and abs(b_tot) < 0.01:
                        continue
                    total_case_dep += c_tot
                    total_bl_dep += b_tot
                    name = get_category_short_name(ce)
                    d = _delta(c_tot, b_tot)
                    dp = _delta_pct(c_tot, b_tot)
                    row = _write_data_row(ws, row, [name, c_tot, b_tot, d, dp],
                                          formats=rf_fmts, delta_cols=[3])

                d = _delta(total_case_dep, total_bl_dep)
                dp = _delta_pct(total_case_dep, total_bl_dep)
                row = _write_data_row(ws, row, ["TOTAL DEPRECIATION", total_case_dep, total_bl_dep, d, dp],
                                      formats=rf_fmts, is_total=True)
                row += 1

    # ---- Section D: M3 Cost of Capital ----
    row = _write_section_header(ws, row, "M3 Cost of Capital", n_cols)
    row = _write_col_headers(ws, row, ["Parameter", "Case", "Baseline", "Delta", ""])

    wacc_case = getattr(case_result.pre_dea, 'wacc_used', None) or 0.0453
    wacc_bl = 0.0453
    wacc_method = getattr(case_result.pre_dea, 'wacc_input_method', 'baseline')
    row = _write_data_row(ws, row, ["WACC (real pre-tax)", wacc_case, wacc_bl,
                                    wacc_case - wacc_bl, None],
                          formats=[None, _FMT_PCT, _FMT_PCT, _FMT_PCT_SIGN, None])
    row = _write_data_row(ws, row, ["Input method", wacc_method, "baseline", "", ""],
                          formats=[None, None, None, None, None])

    # CAPM chain if available
    wacc_inputs = getattr(case_result.pre_dea, 'wacc_inputs', None)
    wacc_derived = getattr(case_result.pre_dea, 'wacc_derived', None)
    if wacc_inputs and isinstance(wacc_inputs, dict):
        for key, val in wacc_inputs.items():
            row = _write_data_row(ws, row, [f"  {key}", val, "", "", ""],
                                  formats=[None, _FMT_RATIO, None, None, None])
    if wacc_derived and isinstance(wacc_derived, dict):
        for key, val in wacc_derived.items():
            row = _write_data_row(ws, row, [f"  {key}", val, "", "", ""],
                                  formats=[None, _FMT_RATIO, None, None, None])

    # Return on assets per year (from SDF/pre_dea, not revenue frame)
    ret_yearly = _get_return_per_year(case_result, baseline_result, user_reid)
    row += 1
    row = _write_col_headers(ws, row, ["Return on assets", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"])
    for year_key in ["2024", "2025", "2026", "2027", "Period"]:
        c_val = ret_yearly[year_key]['case']
        b_val = ret_yearly[year_key]['baseline']
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        is_tot = (year_key == "Period")
        row = _write_data_row(ws, row, [year_key, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3], is_total=is_tot)
    row += 1

    # ---- Section E: M3 Incentives ----
    incentive_details = case_result.post_dea.user_incentive_details
    bl_incentive_details = baseline_result.post_dea.user_incentive_details

    row = _write_section_header(ws, row, "M3 Incentive Adjustments", n_cols)

    # Period totals
    row = _write_col_headers(ws, row, ["Incentive component", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"])
    for label, col_key in [
        ("Quality incentive", COL_QUALITY_INCENTIVE),
        ("Network loss incentive", COL_NETLOSS_INCENTIVE),
        ("Load incentive", COL_LOAD_INCENTIVE),
        ("TOTAL incentive adjustment", COL_INCENTIVE_TOTAL),
    ]:
        c_val = _safe(case_ir, col_key)
        b_val = _safe(bl_ir, col_key)
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        is_tot = (col_key == COL_INCENTIVE_TOTAL)
        row = _write_data_row(ws, row, [label, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3], is_total=is_tot)

    # Per-year breakdown if available
    if incentive_details is not None and not incentive_details.empty:
        row += 1
        year_headers = ["Year", "Quality (tkr)", "Network loss (tkr)", "Load (tkr)", "Total (tkr)"]
        row = _write_col_headers(ws, row, year_headers)
        year_fmts = [None, _FMT_TKR, _FMT_TKR, _FMT_TKR, _FMT_TKR]
        for _, yr_row in incentive_details.iterrows():
            year = yr_row.get("year", "")
            # Incentive detail values are in kr -- convert to tkr
            quality = (yr_row.get("inter_incentive", 0.0) or 0.0) / 1000
            netloss = (yr_row.get("loss_incentive", 0.0) or 0.0) / 1000
            load = (yr_row.get("util_incentive", 0.0) or 0.0) / 1000
            total = (yr_row.get("incentive_total_year", 0.0) or 0.0) / 1000
            row = _write_data_row(ws, row, [year, quality, netloss, load, total],
                                  formats=year_fmts)
    row += 1

    # ---- Section F: M4 Operating Expenditures ----
    row = _write_section_header(ws, row, "M4 Operating Expenditures", n_cols)
    row = _write_col_headers(ws, row, ["Component", "Case (tkr)", "Baseline (tkr)", "Delta (tkr)", "Delta (%)"])

    method = _safe_str(case_ir, COL_METHOD_USED, "OPEX")

    opex_components = [
        ("Controllable OPEX (before eff.)", COL_OPEX_BEFORE, COL_CONTROLLABLE_BEFORE),
        ("OPEX efficiency deduction", COL_OPEX_EFF_DEDUCTION, COL_EFFICIENCY_DEDUCTION),
        ("Non-controllable OPEX", COL_NON_CONTROLLABLE, None),
        ("Flexibility services", COL_FLEXIBILITY, None),
        ("Interruption compensation", COL_INTERRUPTION, None),
        ("State subsidy deduction", COL_STATE_DEDUCTION, None),
    ]

    if method == "TOTEX":
        # Insert CAPEX rows after OPEX eff deduction
        opex_components = [
            ("Controllable OPEX (before eff.)", COL_OPEX_BEFORE, COL_CONTROLLABLE_BEFORE),
            ("OPEX efficiency deduction", COL_OPEX_EFF_DEDUCTION, COL_EFFICIENCY_DEDUCTION),
            ("CAPEX (before eff.)", COL_CAPEX_BEFORE, None),
            ("CAPEX efficiency deduction", COL_CAPEX_EFF_DEDUCTION, None),
            ("Non-controllable OPEX", COL_NON_CONTROLLABLE, None),
            ("Flexibility services", COL_FLEXIBILITY, None),
            ("Interruption compensation", COL_INTERRUPTION, None),
            ("State subsidy deduction", COL_STATE_DEDUCTION, None),
        ]

    for label, col_key, fallback_key in opex_components:
        c_val = _safe(case_ir, col_key, _safe(case_ir, fallback_key) if fallback_key else 0.0)
        b_val = _safe(bl_ir, col_key, _safe(bl_ir, fallback_key) if fallback_key else 0.0)
        d = _delta(c_val, b_val)
        dp = _delta_pct(c_val, b_val)
        row = _write_data_row(ws, row, [label, c_val, b_val, d, dp],
                              formats=rf_fmts, delta_cols=[3])

    row += 1

    # ---- Section G: M5 Efficiency ----
    row = _write_section_header(ws, row, "M5 Efficiency", n_cols)
    row = _write_col_headers(ws, row, ["Metric", "Case", "Baseline", "Delta", ""])

    eff_case = case_result.extraction.efficiency
    eff_bl = baseline_result.extraction.efficiency
    pot_case = case_result.extraction.potential
    pot_bl = baseline_result.extraction.potential
    effkrav_case = case_result.post_dea.user_eff_req_pct
    effkrav_bl = baseline_result.post_dea.user_eff_req_pct

    rank_case, n_total = _compute_rank(case_result.dea.dea_results, user_reid)
    rank_bl, _ = _compute_rank(baseline_result.dea.dea_results, user_reid)

    eff_rows = [
        ("DEA efficiency score", eff_case, eff_bl, _FMT_RATIO),
        ("Rank (1 = best)", rank_case, rank_bl, _FMT_INT),
        ("Companies in sample", n_total, n_total, _FMT_INT),
        ("Potential (raw)", pot_case, pot_bl, _FMT_PCT),
        ("Annual eff. requirement", effkrav_case, effkrav_bl, _FMT_PCT),
        ("Is outlier", "Yes" if case_result.extraction.is_outlier else "No",
         "Yes" if baseline_result.extraction.is_outlier else "No", None),
    ]

    for label, c_val, b_val, fmt in eff_rows:
        d = None
        if isinstance(c_val, (int, float)) and isinstance(b_val, (int, float)):
            d = c_val - b_val if c_val is not None and b_val is not None else None
        fmts = [None, fmt, fmt, fmt, None]
        row = _write_data_row(ws, row, [label, c_val, b_val, d, ""], formats=fmts)
    row += 1

    # ---- Section H: Configuration Changes ----
    modified = _get_modified_config_entries(ui_config)
    if modified:
        row = _write_section_header(ws, row, "Configuration Changes (non-baseline values)", 3)
        row = _write_col_headers(ws, row, ["Module", "Parameter", "Value", "", ""])
        for module, key, value in modified:
            row = _write_data_row(ws, row, [module, key, _format_config_value(value), "", ""])

    # Column widths
    _set_column_widths(ws, [
        (1, _W_LABEL), (2, _W_TKR), (3, _W_TKR), (4, _W_TKR), (5, _W_PCT),
    ])

    # Freeze top rows
    ws.freeze_panes = "A5"


# =============================================================================
# SHEET 2: CAPITAL BASE DETAIL
# =============================================================================

def _build_capital_base_detail_sheet(
    wb: Workbook,
    case_result,
    baseline_result,
    company_name: str,
    user_reid: str,
    case_name: str,
):
    user_id_network = getattr(case_result.pre_dea, 'user_id_network', None)
    if user_id_network is None:
        return

    case_cat, bl_cat = _load_category_data(case_result, baseline_result, user_id_network)
    if case_cat is None and bl_cat is None:
        return

    ws = wb.create_sheet("Capital Base Detail")
    row = _write_header_block(ws, company_name, user_reid, case_name)

    # Flat side-by-side layout: one row per (category, half-year),
    # with Case / Baseline / Delta columns for each component.
    headers = [
        "Category", "Period",
        "NUAV Case (tkr)", "NUAV Baseline (tkr)", "NUAV Δ (tkr)",
        "Dep Case (tkr)", "Dep Baseline (tkr)", "Dep Δ (tkr)",
        "Return Case (tkr)", "Return Baseline (tkr)", "Return Δ (tkr)",
    ]
    fmts = [None, None] + [_FMT_TKR] * 3 + [_FMT_TKR] * 3 + [_FMT_TKR] * 3
    delta_cols = [4, 7, 10]  # 0-based indices for the delta columns

    # Ensure columns exist on both DataFrames
    _value_cols = ['nuav_ord', 'nuav_tail', 'dep_ord', 'dep_tail',
                   'return_ord', 'return_tail']
    for df in [case_cat, bl_cat]:
        if df is not None:
            for col in _value_cols:
                if col not in df.columns:
                    df[col] = 0.0

    # Build a merged index of all (cat_encode, time) pairs
    all_keys = set()
    for df in [case_cat, bl_cat]:
        if df is not None and not df.empty:
            for _, r in df.iterrows():
                all_keys.add((int(r['cat_encode']), int(r.get('time', 0))))

    # Index both DataFrames by (cat_encode, time) for fast lookup
    def _index_df(df):
        lookup = {}
        if df is None or df.empty:
            return lookup
        for _, r in df.iterrows():
            key = (int(r['cat_encode']), int(r.get('time', 0)))
            lookup[key] = r
        return lookup

    case_lookup = _index_df(case_cat)
    bl_lookup = _index_df(bl_cat)

    row = _write_section_header(ws, row, "Capital base — Case vs Baseline", len(headers))
    row = _write_col_headers(ws, row, headers)

    def _total(r, ord_col, tail_col):
        return float(r.get(ord_col, 0) or 0) + float(r.get(tail_col, 0) or 0)

    _zero = pd.Series({c: 0.0 for c in _value_cols})

    for ce, tc in sorted(all_keys):
        cat_name = get_category_short_name(ce)
        period_label = TIMECODE_TO_HALFYEAR.get(tc, str(tc))

        cr = case_lookup.get((ce, tc), _zero)
        br = bl_lookup.get((ce, tc), _zero)

        c_nuav = _total(cr, 'nuav_ord', 'nuav_tail')
        b_nuav = _total(br, 'nuav_ord', 'nuav_tail')
        c_dep = _total(cr, 'dep_ord', 'dep_tail')
        b_dep = _total(br, 'dep_ord', 'dep_tail')
        c_ret = _total(cr, 'return_ord', 'return_tail')
        b_ret = _total(br, 'return_ord', 'return_tail')

        row = _write_data_row(ws, row, [
            cat_name, period_label,
            c_nuav, b_nuav, c_nuav - b_nuav,
            c_dep, b_dep, c_dep - b_dep,
            c_ret, b_ret, c_ret - b_ret,
        ], formats=fmts, delta_cols=delta_cols)

    _set_column_widths(ws, [
        (1, 25), (2, 10),
        (3, 16), (4, 16), (5, 14),
        (6, 16), (7, 16), (8, 14),
        (9, 16), (10, 16), (11, 14),
    ])
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SHEET 3: ALL COMPANIES - REVENUE FRAME
# =============================================================================

def _build_all_companies_revenue_sheet(wb: Workbook, case_result, user_reid: str):
    all_rf = case_result.post_dea.all_revenue_frames
    if all_rf is None or all_rf.empty:
        return

    names = _get_company_names(case_result)

    # Select and rename columns for clean output
    cols_to_include = [
        (COL_REID, "REId"),
    ]

    # Build output DataFrame
    df = all_rf.copy()
    if COL_COMPANY_NAME not in df.columns and names:
        df[COL_COMPANY_NAME] = df[COL_REID].map(names)

    # Select relevant columns in waterfall order (matching Summary Section A)
    output_cols = [COL_REID, COL_COMPANY_NAME]
    rf_cols = [
        COL_CONTROLLABLE_IN_RF, COL_CONTROLLABLE_PERIOD,
        COL_NON_CONTROLLABLE,
        COL_CAPITAL_COST_PERIOD,
        COL_OPEX_EFF_DEDUCTION, COL_CAPEX_EFF_DEDUCTION,
        COL_QUALITY_INCENTIVE, COL_NETLOSS_INCENTIVE, COL_LOAD_INCENTIVE,
        COL_FLEXIBILITY, COL_INTERRUPTION, COL_STATE_DEDUCTION,
        COL_REVENUE_FRAME,
    ]
    selected = [c for c in output_cols + rf_cols if c in df.columns]
    df = df[selected].copy()

    # Rename for readability
    rename = {
        COL_REID: "REId",
        COL_COMPANY_NAME: "Company",
        COL_CONTROLLABLE_IN_RF: "Controllable OPEX (tkr)",
        COL_CONTROLLABLE_PERIOD: "Controllable Period (tkr)",
        COL_NON_CONTROLLABLE: "Non-controllable (tkr)",
        COL_CAPITAL_COST_PERIOD: "Capital Cost (tkr)",
        COL_OPEX_EFF_DEDUCTION: "OPEX Eff. Deduction (tkr)",
        COL_CAPEX_EFF_DEDUCTION: "CAPEX Eff. Deduction (tkr)",
        COL_QUALITY_INCENTIVE: "Quality Incentive (tkr)",
        COL_NETLOSS_INCENTIVE: "Netloss Incentive (tkr)",
        COL_LOAD_INCENTIVE: "Load Incentive (tkr)",
        COL_FLEXIBILITY: "Flexibility (tkr)",
        COL_INTERRUPTION: "Interruption (tkr)",
        COL_STATE_DEDUCTION: "State Deduction (tkr)",
        COL_REVENUE_FRAME: "Revenue Frame (tkr)",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Formats
    col_fmts = {v: _FMT_TKR for k, v in rename.items()
                if "(tkr)" in v and v in df.columns}

    ws = wb.create_sheet("Revenue Frame Decomposition")
    _write_raw_dataframe(ws, df, highlight_reid=user_reid, column_formats=col_fmts)

    # Column widths
    widths = []
    for i, col_name in enumerate(df.columns, 1):
        if col_name == "Company":
            widths.append((i, _W_NAME))
        elif col_name == "REId":
            widths.append((i, _W_REID))
        elif "(tkr)" in col_name:
            widths.append((i, _W_TKR))
        elif "(%)" in col_name:
            widths.append((i, _W_PCT))
        else:
            widths.append((i, _W_SMALL + 4))
    _set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SHEET 4: ALL COMPANIES - DEA
# =============================================================================

def _build_all_companies_dea_sheet(wb: Workbook, case_result, baseline_result, user_reid: str):
    dea = case_result.dea.dea_results
    if dea is None or dea.empty:
        return

    names = _get_company_names(case_result)
    df = dea.copy()
    if COL_COMPANY_NAME not in df.columns and names:
        df[COL_COMPANY_NAME] = df[COL_REID].map(names)

    # Merge case eff req
    all_eff = case_result.post_dea.all_eff_reqs
    if all_eff is not None and COL_EFF_REQ_ANNUAL not in df.columns:
        df = df.merge(all_eff[[COL_REID, COL_EFF_REQ_ANNUAL]], on=COL_REID, how="left")

    # Merge baseline efficiency + eff req for comparison
    bl_dea = baseline_result.dea.dea_results
    if bl_dea is not None and not bl_dea.empty:
        bl_cols = [COL_REID]
        bl_rename = {}
        if COL_DEA_EFFICIENCY in bl_dea.columns:
            bl_cols.append(COL_DEA_EFFICIENCY)
            bl_rename[COL_DEA_EFFICIENCY] = "bl_dea_efficiency"
        if COL_DEA_POTENTIAL in bl_dea.columns:
            bl_cols.append(COL_DEA_POTENTIAL)
            bl_rename[COL_DEA_POTENTIAL] = "bl_potential"
        df = df.merge(
            bl_dea[bl_cols].rename(columns=bl_rename),
            on=COL_REID, how="left",
        )
    bl_eff_reqs = baseline_result.post_dea.all_eff_reqs
    if bl_eff_reqs is not None and COL_EFF_REQ_ANNUAL in bl_eff_reqs.columns:
        df = df.merge(
            bl_eff_reqs[[COL_REID, COL_EFF_REQ_ANNUAL]].rename(
                columns={COL_EFF_REQ_ANNUAL: "bl_eff_req_annual"}
            ),
            on=COL_REID, how="left",
        )

    # Merge DEA input variables from pre_dea (costs + outputs used in DEA)
    df_all = case_result.pre_dea.df_all_companies
    dea_input_cols = [COL_CAPITAL_COST_2024, COL_CONTROLLABLE_AVG,
                      COL_CU, COL_MW, COL_NS, COL_MWH_LOW, COL_MWH_HIGH]
    merge_cols = [c for c in dea_input_cols if c in df_all.columns]
    if merge_cols and COL_REID in df_all.columns:
        df = df.merge(df_all[[COL_REID] + merge_cols], on=COL_REID, how="left")

    output_cols = [COL_REID, COL_COMPANY_NAME,
                   COL_DEA_EFFICIENCY, "bl_dea_efficiency",
                   COL_DEA_SUPER_EFF,
                   COL_DEA_POTENTIAL, "bl_potential",
                   COL_IS_OUTLIER,
                   COL_EFF_REQ_ANNUAL, "bl_eff_req_annual",
                   COL_CAPITAL_COST_2024, COL_CONTROLLABLE_AVG,
                   COL_CU, COL_MW, COL_NS, COL_MWH_LOW, COL_MWH_HIGH]
    selected = [c for c in output_cols if c in df.columns]
    df = df[selected].copy()

    rename = {
        COL_REID: "REId",
        COL_COMPANY_NAME: "Company",
        COL_DEA_EFFICIENCY: "DEA Efficiency (Case)",
        "bl_dea_efficiency": "DEA Efficiency (Baseline)",
        COL_DEA_SUPER_EFF: "Super-Efficiency",
        COL_DEA_POTENTIAL: "Potential (Case)",
        "bl_potential": "Potential (Baseline)",
        COL_IS_OUTLIER: "Outlier",
        COL_EFF_REQ_ANNUAL: "Eff. Req (Case)",
        "bl_eff_req_annual": "Eff. Req (Baseline)",
        COL_CAPITAL_COST_2024: "Cap Cost 2024 (tkr)",
        COL_CONTROLLABLE_AVG: "Controllable Avg (tkr)",
        COL_CU: "CU",
        COL_MW: "MW",
        COL_NS: "NS",
        COL_MWH_LOW: "MWhl",
        COL_MWH_HIGH: "MWhh",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    col_fmts = {
        "DEA Efficiency (Case)": _FMT_RATIO,
        "DEA Efficiency (Baseline)": _FMT_RATIO,
        "Super-Efficiency": _FMT_RATIO,
        "Potential (Case)": _FMT_PCT,
        "Potential (Baseline)": _FMT_PCT,
        "Eff. Req (Case)": _FMT_PCT,
        "Eff. Req (Baseline)": _FMT_PCT,
        "Cap Cost 2024 (tkr)": _FMT_TKR,
        "Controllable Avg (tkr)": _FMT_TKR,
        "CU": _FMT_TKR,
        "MW": '#,##0.0',
        "NS": _FMT_TKR,
        "MWhl": _FMT_TKR,
        "MWhh": _FMT_TKR,
    }

    ws = wb.create_sheet("DEA")
    _write_raw_dataframe(ws, df, highlight_reid=user_reid, column_formats=col_fmts)

    widths = [(1, _W_REID), (2, _W_NAME)]
    for i, col_name in enumerate(df.columns[2:], 3):
        if "(tkr)" in col_name or col_name in ("CU", "MW", "NS", "MWhl", "MWhh"):
            widths.append((i, _W_TKR))
        else:
            widths.append((i, _W_PCT))
    _set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SHEET 5: ALL COMPANIES - COSTS
# =============================================================================

def _build_all_companies_costs_sheet(wb: Workbook, case_result, user_reid: str):
    df_all = case_result.pre_dea.df_all_companies
    if df_all is None or df_all.empty:
        return

    output_cols = [
        COL_REID, COL_COMPANY_NAME,
        COL_CAPITAL_COST_2024, COL_CAPITAL_COST_2025, COL_CAPITAL_COST_2026, COL_CAPITAL_COST_2027,
        COL_CAPITAL_COST_PERIOD,
        COL_CONTROLLABLE_AVG,
        COL_TOTEX,
        COL_DEPRECIATION_PERIOD, COL_RETURN_PERIOD,
    ]
    selected = [c for c in output_cols if c in df_all.columns]
    df = df_all[selected].copy()

    rename = {
        COL_REID: "REId",
        COL_COMPANY_NAME: "Company",
        COL_CAPITAL_COST_2024: "Cap Cost 2024 (tkr)",
        COL_CAPITAL_COST_2025: "Cap Cost 2025 (tkr)",
        COL_CAPITAL_COST_2026: "Cap Cost 2026 (tkr)",
        COL_CAPITAL_COST_2027: "Cap Cost 2027 (tkr)",
        COL_CAPITAL_COST_PERIOD: "Cap Cost Period (tkr)",
        COL_CONTROLLABLE_AVG: "Controllable Avg (tkr)",
        COL_TOTEX: "TOTEX 1st Year (tkr)",
        COL_DEPRECIATION_PERIOD: "Depreciation Period (tkr)",
        COL_RETURN_PERIOD: "Return Period (tkr)",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    col_fmts = {v: _FMT_TKR for k, v in rename.items() if "(tkr)" in v and v in df.columns}

    ws = wb.create_sheet("Costs")
    _write_raw_dataframe(ws, df, highlight_reid=user_reid, column_formats=col_fmts)

    widths = [(1, _W_REID), (2, _W_NAME)]
    for i in range(3, len(df.columns) + 1):
        widths.append((i, _W_TKR))
    _set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SHEET 6: ALL COMPANIES - INCENTIVES
# =============================================================================

def _build_all_companies_incentives_sheet(wb: Workbook, case_result, user_reid: str):
    all_inc = case_result.post_dea.all_incentives
    if all_inc is None or all_inc.empty:
        return

    names = _get_company_names(case_result)
    df = all_inc.copy()
    if COL_COMPANY_NAME not in df.columns and names:
        df[COL_COMPANY_NAME] = df[COL_REID].map(names)

    # Merge OPEX/CAPEX efficiency deductions from revenue frames
    all_rf = case_result.post_dea.all_revenue_frames
    if all_rf is not None and not all_rf.empty:
        eff_cols = [c for c in [COL_OPEX_EFF_DEDUCTION, COL_CAPEX_EFF_DEDUCTION]
                    if c in all_rf.columns]
        if eff_cols and COL_REID in all_rf.columns:
            df = df.merge(all_rf[[COL_REID] + eff_cols], on=COL_REID, how="left")

    output_cols = [COL_REID, COL_COMPANY_NAME,
                   COL_OPEX_EFF_DEDUCTION, COL_CAPEX_EFF_DEDUCTION,
                   COL_QUALITY_INCENTIVE, COL_NETLOSS_INCENTIVE,
                   COL_LOAD_INCENTIVE, COL_MISSING_INCENTIVE]
    selected = [c for c in output_cols if c in df.columns]
    df = df[selected].copy()

    rename = {
        COL_REID: "REId",
        COL_COMPANY_NAME: "Company",
        COL_OPEX_EFF_DEDUCTION: "OPEX Eff. Deduction (tkr)",
        COL_CAPEX_EFF_DEDUCTION: "CAPEX Eff. Deduction (tkr)",
        COL_QUALITY_INCENTIVE: "Quality Incentive (tkr)",
        COL_NETLOSS_INCENTIVE: "Netloss Incentive (tkr)",
        COL_LOAD_INCENTIVE: "Load Incentive (tkr)",
        COL_MISSING_INCENTIVE: "Missing Data",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    col_fmts = {v: _FMT_TKR for k, v in rename.items() if "(tkr)" in v and v in df.columns}

    ws = wb.create_sheet("Incentives")
    _write_raw_dataframe(ws, df, highlight_reid=user_reid, column_formats=col_fmts)

    widths = [(1, _W_REID), (2, _W_NAME)]
    for i in range(3, len(df.columns) + 1):
        widths.append((i, _W_TKR))
    _set_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# SHEET 7: CONFIGURATION
# =============================================================================

def _build_config_sheet(
    wb: Workbook,
    case_result,
    ui_config: Dict[str, Any],
):
    ws = wb.create_sheet("Configuration")

    row = 1
    row = _write_section_header(ws, row, "Pipeline Metadata", 3)
    row = _write_col_headers(ws, row, ["Property", "Value", ""])

    metadata = [
        ("Capbase source", getattr(case_result.pre_dea, 'capbase_source', 'baseline')),
        ("CAPEX method", getattr(case_result.pre_dea, 'capex_method', 'baseline')),
        ("CAPEX modified", "Yes" if getattr(case_result.pre_dea, 'capex_modified', False) else "No"),
        ("OPEX modified", "Yes" if getattr(case_result.pre_dea, 'opex_modified', False) else "No"),
        ("WACC input method", getattr(case_result.pre_dea, 'wacc_input_method', 'baseline')),
        ("WACC used", getattr(case_result.pre_dea, 'wacc_used', 0.0453)),
        ("DEA method", getattr(case_result.dea, 'dea_method', 'baseline')),
        ("DEA executed", "Yes" if getattr(case_result.dea, 'dea_executed', False) else "No"),
    ]

    for key, val in metadata:
        fmt = _FMT_PCT if isinstance(val, float) else None
        row = _write_data_row(ws, row, [key, val, ""], formats=[None, fmt, None])

    row += 1
    row = _write_section_header(ws, row, "Full UI Configuration", 3)
    row = _write_col_headers(ws, row, ["Module.Parameter", "Value", ""])

    for section, params in ui_config.items():
        if isinstance(params, dict):
            for key, val in params.items():
                row = _write_data_row(ws, row,
                                      [f"{section}.{key}", _format_config_value(val), ""])
        else:
            row = _write_data_row(ws, row,
                                  [section, _format_config_value(params), ""])

    _set_column_widths(ws, [(1, 40), (2, 50), (3, 10)])


# =============================================================================
# CONFIG HELPERS
# =============================================================================

def _get_modified_config_entries(ui_config: Dict[str, Any]) -> List[Tuple[str, str, Any]]:
    """Return list of (module, key, value) for non-None/non-default entries."""
    modified = []
    for section, params in ui_config.items():
        if not isinstance(params, dict):
            continue
        for key, val in params.items():
            if val is None:
                continue
            # Skip bytes (file uploads) and empty dicts/lists
            if isinstance(val, bytes):
                modified.append((section, key, f"<file: {len(val)} bytes>"))
            elif isinstance(val, dict) and not val:
                continue
            elif isinstance(val, list) and not val:
                continue
            else:
                modified.append((section, key, val))
    return modified


def _format_config_value(val) -> str:
    """Format a config value for display in the configuration sheet."""
    if val is None:
        return "(baseline)"
    if isinstance(val, bytes):
        return f"<file: {len(val)} bytes>"
    if isinstance(val, dict):
        if not val:
            return "(empty)"
        return str(val)
    if isinstance(val, list):
        if not val:
            return "(empty)"
        return ", ".join(str(v) for v in val)
    if isinstance(val, float):
        return f"{val:.6g}"
    return str(val)


# =============================================================================
# PUBLIC API
# =============================================================================

def create_case_export(
    user_reid: str,
    company_name: str,
    baseline_result,
    case_result,
    ui_config: Dict[str, Any],
    case_name: str = "",
) -> BytesIO:
    """Create professional Excel export from pipeline results.

    Returns BytesIO containing the .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary (polished)
    _build_summary_sheet(wb, case_result, baseline_result, ui_config,
                         company_name, user_reid, case_name)

    # Sheet 2: Capital Base Detail (half-year per category)
    _build_capital_base_detail_sheet(wb, case_result, baseline_result,
                                     company_name, user_reid, case_name)

    # Sheet 3-6: Raw data (all 148 companies)
    _build_all_companies_revenue_sheet(wb, case_result, user_reid)
    _build_all_companies_dea_sheet(wb, case_result, baseline_result, user_reid)
    _build_all_companies_costs_sheet(wb, case_result, user_reid)
    _build_all_companies_incentives_sheet(wb, case_result, user_reid)

    # Sheet 7: Configuration
    _build_config_sheet(wb, case_result, ui_config)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_export_filename(user_reid: str) -> str:
    """Generate filename with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"regumetrica_{user_reid}_{timestamp}.xlsx"
