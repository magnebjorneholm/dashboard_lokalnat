#!/usr/bin/env python3
"""
build_lokalnat_data.py
======================
Reproducible ETL that turns the raw Energimarknadsinspektionen (Ei) Excel
appendices in ``raw/`` into tidy, analysis-ready datasets in ``lokalnat/``.

What it does
------------
* Reads every workbook in ``raw/`` (report Ei R2026:05, appendices 1-8).
* Reshapes the wide, multi-row-header sheets into **tidy long format**
  (one observation per row).
* Keeps **local grid (lokalnät) only** - regional grid (regionnät) and
  transmission grid (transmissionsnät) rows are dropped, per the data owner's
  request. Filtering is done on the ``Nätnivå`` column where present, otherwise
  on the accounting-unit id prefix (REL = local, RER = regional, RET =
  transmission).
* Writes one Parquet **and** one CSV file per logical dataset, plus reference
  tables (metric dictionary, company network-level map) and a JSON manifest.

Run:  ``uv run python build_lokalnat_data.py``   (from this directory)
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
OUT = HERE / "lokalnat"
PARQUET = OUT / "parquet"
CSV = OUT / "csv"
REF = OUT / "reference"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
CODE_RE = re.compile(r"^(RR|BR|IB|TU)\d+$")          # Ei chart-of-accounts codes
NULLS = {None, "-", "_", "", "–"}                    # blank markers used in the files
ID_LABELS = {"Företag", "Elnätsföretag", "Redovisningsenhet", "Org.nr", "Nätnivå"}

LEVEL_BY_PREFIX = {"REL": "Lokalnät", "RER": "Regionnät", "RET": "Transmissionsnät"}


def raw_file(bilaga: int) -> Path:
    """Locate a raw workbook by its appendix (bilaga) number."""
    hits = sorted(RAW.glob(f"*bilaga-{bilaga}-*.xlsx"))
    if not hits:
        raise FileNotFoundError(f"No raw workbook for bilaga {bilaga} in {RAW}")
    return hits[0]


def is_year(v) -> bool:
    if isinstance(v, int):
        return 2000 <= v <= 2035
    if isinstance(v, str) and v.strip().isdigit():
        return 2000 <= int(v.strip()) <= 2035
    return False


def clean_num(v):
    """Coerce a cell to float, mapping blank markers to NaN."""
    if v in NULLS:
        return pd.NA
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if s in {"", "-", "–"}:
            return pd.NA
        try:
            return float(s)
        except ValueError:
            return pd.NA
    return v


def clean_str(v):
    if v in NULLS:
        return None
    return str(v).strip()


def level_from_reid(reid: str | None):
    if not reid:
        return None
    return LEVEL_BY_PREFIX.get(reid[:3])


def infer_unit(name: str | None, default: str) -> str:
    n = (name or "").lower()
    if "kilometer" in n:
        return "km"
    if "mva" in n:
        return "MVA"
    if "(mw)" in n:
        return "MW"
    if n.startswith("antal") or "abonnemang" in n:
        return "count"
    if "i förhållande till" in n or n.startswith("andel"):
        return "ratio"
    if any(k in n for k in ("soliditet", "marginal", "avkastning på", "skuldsättningsgrad")):
        return "ratio"
    return default


# ---------------------------------------------------------------------------
# Generic parser for the "wide, years across columns" sheets
# ---------------------------------------------------------------------------
def parse_yearly_sheet(ws, dataset: str, default_unit: str, source: str) -> pd.DataFrame:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # locate the row that holds the id labels + the year columns
    year_row_idx = None
    for i, r in enumerate(rows[:6]):
        if sum(1 for c in r if is_year(c)) >= 2:
            year_row_idx = i
            break
    if year_row_idx is None:
        raise ValueError(f"No year row found in sheet {ws.title!r}")
    year_row = rows[year_row_idx]

    # id columns (located by label, so column order does not matter)
    id_cols = {label: j for j, c in enumerate(year_row)
               for label in [c] if label in ID_LABELS}
    has_reid = "Redovisningsenhet" in id_cols
    granularity = "accounting_unit" if has_reid else "company"

    year_cols = [j for j, c in enumerate(year_row) if is_year(c)]

    # classify the metadata rows above the year row into code-row / name-row
    meta_rows = rows[:year_row_idx]
    code_row = name_row = None
    best_codes = best_names = -1
    for mr in meta_rows:
        vals = [mr[j] for j in year_cols if j < len(mr)]
        n_codes = sum(1 for v in vals if isinstance(v, str) and CODE_RE.match(v.strip()))
        n_names = sum(1 for v in vals if v not in NULLS and not (isinstance(v, str) and CODE_RE.match(v.strip())))
        if n_codes > best_codes and n_codes > 0:
            best_codes, code_row = n_codes, mr
        if n_names > best_names:
            best_names, name_row = n_names, mr
    if code_row is name_row:        # a single metadata row that holds names
        code_row = None

    records = []
    for r in rows[year_row_idx + 1:]:
        company = clean_str(r[id_cols["Företag"]] if "Företag" in id_cols
                            else r[id_cols.get("Elnätsföretag", -1)])
        org_nr = clean_str(r[id_cols["Org.nr"]]) if "Org.nr" in id_cols and id_cols["Org.nr"] < len(r) else None
        reid = clean_str(r[id_cols["Redovisningsenhet"]]) if has_reid else None
        net = clean_str(r[id_cols["Nätnivå"]]) if "Nätnivå" in id_cols else None
        if not company and not org_nr and not reid:
            continue                                  # blank / footnote row
        if company and company.startswith("*"):
            continue
        for j in year_cols:
            if j >= len(r):
                continue
            code = code_row[j] if code_row and j < len(code_row) else None
            code = code.strip() if isinstance(code, str) and CODE_RE.match(code.strip()) else None
            name = name_row[j] if name_row and j < len(name_row) else None
            name = None if name in NULLS else str(name).replace("\n", " ").strip()
            if name is None and code is None:
                name = dataset
            records.append({
                "dataset": dataset,
                "granularity": granularity,
                "company": company,
                "org_nr": org_nr,
                "re_id": reid,
                "network_level": net,
                "metric_code": code,
                "metric_name": name if name is not None else code,
                "year": int(year_row[j]),
                "value": clean_num(r[j]),
                "unit": None,
                "source_file": source,
            })

    df = pd.DataFrame.from_records(records)

    # derive network level for accounting units that lack a Nätnivå column
    if has_reid and "Nätnivå" not in id_cols:
        df["network_level"] = df["re_id"].map(level_from_reid)

    df["unit"] = df["metric_name"].map(lambda n: infer_unit(n, default_unit))
    return df


# ---------------------------------------------------------------------------
# Investments (bilaga 2) - values are period totals, not per year
# ---------------------------------------------------------------------------
INV_CATEGORY = {
    "totala": "total",
    "reinvesteringar": "reinvestment",
    "nyinvesteringar": "new_investment",
}


def parse_investments_sheet(ws, source: str) -> pd.DataFrame:
    title = ws.title  # e.g. "Utfall 2020-2023 per Re-ID"
    basis = "outcome" if title.startswith("Utfall") else "forecast"
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = rows[1]                       # real column names live on row 2
    id_cols = {c: j for j, c in enumerate(header) if c in ID_LABELS}
    has_reid = "Redovisningsenhet" in id_cols
    granularity = "accounting_unit" if has_reid else "company"

    # value columns: everything that is not an id column
    val_cols = {}
    window = None
    for j, c in enumerate(header):
        if c in ID_LABELS or c in NULLS or c is None:
            continue
        cl = str(c).lower()
        cat = next((v for k, v in INV_CATEGORY.items() if k in cl), None)
        if cat:
            val_cols[j] = cat
        m = re.search(r"(\d{4})\s*H\d\s*-\s*(\d{4})\s*H\d", str(c))
        if m and window is None:
            window = f"{m.group(1)}H2-{m.group(2)}H1"

    records = []
    for r in rows[2:]:
        company = clean_str(r[id_cols["Elnätsföretag"]])
        org_nr = clean_str(r[id_cols["Org.nr"]]) if "Org.nr" in id_cols else None
        reid = clean_str(r[id_cols["Redovisningsenhet"]]) if has_reid else None
        net = clean_str(r[id_cols["Nätnivå"]]) if "Nätnivå" in id_cols else None
        if not company:
            continue
        for j, cat in val_cols.items():
            if j >= len(r):
                continue
            records.append({
                "dataset": "investments",
                "granularity": granularity,
                "company": company,
                "org_nr": org_nr,
                "re_id": reid,
                "network_level": net,
                "basis": basis,
                "period_window": window,
                "period_start": int(window[:4]) if window else None,
                "period_end": int(window.split("-")[1][:4]) if window else None,
                "investment_category": cat,
                "value": clean_num(r[j]),
                "unit": "tkr (2022 price level)",
                "source_file": source,
            })
    return pd.DataFrame.from_records(records)


# ---------------------------------------------------------------------------
# Accounting-unit changes (bilaga 1) - year-paired matrix
# ---------------------------------------------------------------------------
def parse_changes_sheet(ws, source: str) -> pd.DataFrame:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    year_idx = next(i for i, r in enumerate(rows) if sum(1 for c in r if is_year(c)) >= 2)
    years = rows[year_idx]
    # each (re_id, name) pair sits in two consecutive columns under a year
    pairs = []
    j = 0
    while j < len(years):
        if is_year(years[j]):
            pairs.append((int(years[j]), j, j + 1))
            j += 2
        else:
            j += 1

    records = []
    for track, r in enumerate(rows[year_idx + 2:]):       # +2: skip the sub-header row
        if r and isinstance(r[0], str) and r[0].startswith("*"):
            continue
        for yr, jc, jn in pairs:
            reid = clean_str(r[jc]) if jc < len(r) else None
            name = clean_str(r[jn]) if jn < len(r) else None
            if not reid and not name:
                continue
            records.append({"track_id": track, "year": yr, "re_id": reid, "company": name})

    df = pd.DataFrame.from_records(records)

    # reconstruct the change type (the source encodes it only via cell colour)
    def classify(g):
        g = g.sort_values("year")
        prev_reid = prev_name = None
        out = []
        first = True
        for _, row in g.iterrows():
            if first and row.year > df.year.min():
                ct = "new_accounting_unit"
            elif prev_reid is not None and row.re_id != prev_reid:
                ct = "reid_change"
            elif prev_name is not None and row.company != prev_name:
                ct = "name_change"
            else:
                ct = "no_change"
            out.append(ct)
            prev_reid, prev_name, first = row.re_id, row.company, False
        g = g.copy()
        g["change_type"] = out
        return g

    df = (df.groupby("track_id", group_keys=False)[df.columns.tolist()]
            .apply(classify)
            .reset_index(drop=True))
    df["network_level"] = df["re_id"].map(level_from_reid)
    return df


# ---------------------------------------------------------------------------
# Lokalnät filter
# ---------------------------------------------------------------------------
def filter_local(df: pd.DataFrame) -> pd.DataFrame:
    if "network_level" in df.columns and df["network_level"].notna().any():
        keep = df["network_level"].isna() | df["network_level"].eq("Lokalnät")
        return df[keep].reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Sheet -> dataset routing
# ---------------------------------------------------------------------------
YEARLY = {
    8: {
        "Avkastning per org. nr": ("return_on_capital", "tkr"),
        "EBIT,EBITA, RFP, Avkastning IR": ("return_on_capital", "tkr"),
        "RR, EBITA, RFP per Re-ID": ("return_on_capital", "tkr"),
        "Sysselsatt kapital per Re-ID": ("capital_employed_components", "tkr"),
    },
    6: {
        "Per org.nr": ("group_contributions", "ratio"),
        "Per Re-ID": ("group_contributions", "ratio"),
    },
    4: {
        "Nyckeltal per org.nr": ("key_figures", "ratio"),
        "Underliggande data per Re-ID": ("key_figures_underlying", "tkr"),
    },
    7: {
        "Per org. nr": ("income_statement_items", "tkr"),
        "Per Re-ID": ("income_statement_items", "tkr"),
    },
    5: {
        "Interna, externa lån - org. nr": ("interest_and_loans", "tkr"),
    },
    3: {  # technical indicators: every sheet is one metric x one granularity
        s: ("technical_indicators", "count")
        for s in (
            "Ledningslängd per Re-ID", "Ledningslängd per org.nr",
            "Transformatoreffekt per Re-ID", "Transformatoreffekt per org.nr",
            "Abonnerad effekt per Re-ID", "Abonnerad effekt per org.nr",
            "Uttagspunkt per Re-ID", "Uttagspunkt per org.nr",
            "Inmatningspunkt per Re-ID", "Inmatningspunkt per org.nr",
        )
    },
}

YEARLY_COLS = ["dataset", "granularity", "company", "org_nr", "re_id",
               "network_level", "metric_code", "metric_name", "year",
               "value", "unit", "source_file"]


def build_company_network_levels() -> pd.DataFrame:
    """org_nr -> set of network levels it operates (from sheets carrying Nätnivå)."""
    frames = []
    for bilaga in (2, 3):
        wb = openpyxl.load_workbook(raw_file(bilaga), read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(max_row=3, values_only=True)]
            hdr_idx = next((i for i, r in enumerate(rows)
                            if "Nätnivå" in r and ("Org.nr" in r)), None)
            if hdr_idx is None:
                continue
            hdr = rows[hdr_idx]
            jc = hdr.index("Org.nr"); jn = hdr.index("Nätnivå")
            jcomp = hdr.index("Elnätsföretag") if "Elnätsföretag" in hdr else None
            for r in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
                org = clean_str(r[jc]) if jc < len(r) else None
                lvl = clean_str(r[jn]) if jn < len(r) else None
                comp = clean_str(r[jcomp]) if jcomp is not None and jcomp < len(r) else None
                if org and lvl:
                    frames.append((org, comp, lvl))
        wb.close()
    raw = pd.DataFrame(frames, columns=["org_nr", "company", "network_level"]).drop_duplicates()
    g = raw.groupby("org_nr").agg(
        company=("company", lambda s: next((x for x in s if x), None)),
        network_levels=("network_level", lambda s: ", ".join(sorted(set(s)))),
    ).reset_index()
    g["has_local"] = g["network_levels"].str.contains("Lokalnät")
    g["has_regional"] = g["network_levels"].str.contains("Regionnät")
    g["has_transmission"] = g["network_levels"].str.contains("Transmissionsnät")
    return g


def write(df: pd.DataFrame, name: str):
    df.to_parquet(PARQUET / f"{name}.parquet", index=False)
    df.to_csv(CSV / f"{name}.csv", index=False)
    return {"rows": int(len(df)), "columns": list(df.columns)}


def main():
    for d in (PARQUET, CSV, REF):
        d.mkdir(parents=True, exist_ok=True)

    manifest: dict = {"source_report": "Energimarknadsinspektionen R2026:05",
                      "scope": "Lokalnät only (regionnät & transmissionsnät removed)",
                      "datasets": {}}

    # --- company network-level reference (built first; used to flag company rows)
    cnl = build_company_network_levels()
    cnl.to_parquet(REF / "company_network_levels.parquet", index=False)
    cnl.to_csv(REF / "company_network_levels.csv", index=False)
    nonlocal_orgs = set(cnl.loc[cnl["has_regional"] | cnl["has_transmission"], "org_nr"])

    # --- yearly datasets
    collected: dict[str, list[pd.DataFrame]] = {}
    for bilaga, sheets in YEARLY.items():
        wb = openpyxl.load_workbook(raw_file(bilaga), read_only=True, data_only=True)
        src = raw_file(bilaga).name
        for sheet, (dataset, unit) in sheets.items():
            df = parse_yearly_sheet(wb[sheet], dataset, unit, src)
            collected.setdefault(dataset, []).append(df)
        wb.close()

    all_long = []
    for dataset, parts in collected.items():
        df = pd.concat(parts, ignore_index=True)[YEARLY_COLS]
        df = filter_local(df)
        # flag company-level rows whose org_nr also runs a non-local grid
        df["company_has_nonlocal_units"] = (
            (df["granularity"] == "company") & df["org_nr"].isin(nonlocal_orgs)
        )
        df = df.dropna(subset=["value"]).reset_index(drop=True)
        info = write(df, dataset)
        info["years"] = [int(df["year"].min()), int(df["year"].max())]
        info["metrics"] = int(df["metric_name"].nunique())
        info["granularities"] = sorted(df["granularity"].unique())
        manifest["datasets"][dataset] = info
        all_long.append(df)

    # --- combined long table across all yearly datasets
    combined = pd.concat(all_long, ignore_index=True)
    info = write(combined, "all_yearly_long")
    info["years"] = [int(combined["year"].min()), int(combined["year"].max())]
    manifest["datasets"]["all_yearly_long"] = info

    # --- investments (bilaga 2)
    wb = openpyxl.load_workbook(raw_file(2), read_only=True, data_only=True)
    src = raw_file(2).name
    inv = pd.concat([parse_investments_sheet(ws, src) for ws in wb.worksheets],
                    ignore_index=True)
    wb.close()
    inv = filter_local(inv).dropna(subset=["value"]).reset_index(drop=True)
    info = write(inv, "investments")
    info["periods"] = sorted(inv["period_window"].dropna().unique().tolist())
    info["basis"] = sorted(inv["basis"].unique())
    manifest["datasets"]["investments"] = info

    # --- accounting-unit changes (bilaga 1)
    wb = openpyxl.load_workbook(raw_file(1), read_only=True, data_only=True)
    chg = parse_changes_sheet(wb.worksheets[0], raw_file(1).name)
    wb.close()
    chg = chg[chg["re_id"].str.startswith("REL", na=False)].reset_index(drop=True)
    info = write(chg, "accounting_unit_changes")
    info["years"] = [int(chg["year"].min()), int(chg["year"].max())]
    manifest["datasets"]["accounting_unit_changes"] = info

    # --- metric dictionary
    md = (combined[["dataset", "metric_code", "metric_name", "unit"]]
          .drop_duplicates().sort_values(["dataset", "metric_name"]).reset_index(drop=True))
    md.to_parquet(REF / "metric_dictionary.parquet", index=False)
    md.to_csv(REF / "metric_dictionary.csv", index=False)

    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    print(json.dumps(manifest, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
