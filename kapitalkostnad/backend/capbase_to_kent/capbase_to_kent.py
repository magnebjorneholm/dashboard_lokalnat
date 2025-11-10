"""
capbase_to_kent.py - Reverse engineering från capbase_a till KENT Excel-format

Skapar en KENT Excel-fil från capbase_a.parquet som vid körning genom
capbase_prep.py genererar samma capbase_a tillbaka (round-trip).
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Optional, Tuple
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def time_to_year_halfyear(time_value: float) -> Tuple[int, int]:
    """Konverterar time till år och halvår."""
    if pd.isna(time_value):
        return None, None
    year = int((time_value - 1) // 2 + 1910)
    halfyear = int((time_value - 1) % 2 + 1)
    return year, halfyear


def time_to_year(time_value: float) -> Optional[int]:
    """Konverterar time till år (använder halvår 1 per default)."""
    year, _ = time_to_year_halfyear(time_value)
    return year


def load_uppslagsvarden(uppslagsvarden_path: str) -> pd.DataFrame:
    """Läser uppslagsvärden från KENT Excel-mal."""
    df = pd.read_excel(
        uppslagsvarden_path,
        sheet_name='Uppslagsvärden',
        header=0
    )
    
    # Ta endast relevanta kolumner
    df = df[['Anläggningskategori', 'Kod', 'Typ av anläggning', 
             'Teknisk specifikation', 'Spänning kV', 'Normvärde 2022 (SEK)']].copy()
    
    # Rensa tomma rader
    df = df[df['Kod'].notna()].copy()
    
    return df


def create_normvarde_sheet(capbase_df: pd.DataFrame, uppslagsvarden_df: pd.DataFrame) -> pd.DataFrame:
    """Skapar Normvärde-ark från capbase_a."""
    
    # Filtrera normvärde-komponenter
    normvarde = capbase_df[capbase_df['metod'] == 'normvärde'].copy()
    
    if normvarde.empty:
        return pd.DataFrame()
    
    # Skapa KENT-format
    kent_normvarde = pd.DataFrame()
    
    # Kod - direkt från id_comptype
    kent_normvarde['Kod'] = normvarde['id_comptype']
    
    # Antal - från count_comp
    kent_normvarde['Antal'] = normvarde['count_comp']
    
    # Rådighet
    kent_normvarde['Rådighet'] = normvarde['owned'].map({1: 'Ägd', 0: 'Hyrd/Leasad'})
    
    # Ursprungligen tagen i bruk
    kent_normvarde['Ursprungligen tagen i bruk'] = normvarde['time_from'].apply(time_to_year)
    
    # Hantera time_from_missing
    mask_missing = normvarde['time_from_missing'] == 1
    if mask_missing.any():
        kent_normvarde.loc[mask_missing, 'År saknas (Ja eller blank)'] = 'Ja'
        kent_normvarde.loc[mask_missing, 'Ursprungligen tagen i bruk'] = ''
    
    # Anmärkning (tom kolumn)
    kent_normvarde['Anmärkning'] = ''
    
    return kent_normvarde


def create_ovriga_metoder_sheet(capbase_df: pd.DataFrame) -> pd.DataFrame:
    """Skapar Övriga värderingsmetoder-ark från capbase_a."""
    
    # Filtrera övriga metoder
    ovriga = capbase_df[
        capbase_df['metod'].isin(['anskaffningsvärde', 'bokförtvärde', 'annatskäligtvärde'])
    ].copy()
    
    if ovriga.empty:
        return pd.DataFrame()
    
    # Skapa KENT-format
    kent_ovriga = pd.DataFrame()
    
    # Metodmarkeringar
    kent_ovriga['Ansk'] = ovriga['metod'].apply(lambda x: 'x' if x == 'anskaffningsvärde' else '')
    kent_ovriga['Bokf'] = ovriga['metod'].apply(lambda x: 'x' if x == 'bokförtvärde' else '')
    kent_ovriga['Annat'] = ovriga['metod'].apply(lambda x: 'x' if x == 'annatskäligtvärde' else '')
    
    # Anl.kategori och Typ av anläggning
    kent_ovriga['Anl.kategori'] = ovriga['cat']
    kent_ovriga['Typ av anläggning'] = ovriga['subcat']
    
    # Antal
    kent_ovriga['Antal'] = ovriga['count_comp']
    
    # Ursprungligen tagen i bruk
    kent_ovriga['Ursprungligen tagen i bruk'] = ovriga['time_from'].apply(time_to_year)
    
    # Rådighet
    kent_ovriga['Rådighet'] = ovriga['owned'].map({1: 'Ägd', 0: 'Hyrd/Leasad'})
    
    # NUAV 2022 (kr)
    kent_ovriga['NUAV 2022 (kr)'] = ovriga['nuav_2022']
    
    # Hantera time_from_missing
    mask_missing = ovriga['time_from_missing'] == 1
    if mask_missing.any():
        kent_ovriga.loc[mask_missing, 'År saknas (Ja eller blank)'] = 'Ja'
        kent_ovriga.loc[mask_missing, 'Ursprungligen tagen i bruk'] = ''
    
    # Anmärkning
    kent_ovriga['Anmärkning'] = ''
    
    return kent_ovriga


def create_investeringar_sheet(capbase_df: pd.DataFrame) -> pd.DataFrame:
    """Skapar Investeringar_Utrangeringar-ark från capbase_a."""
    
    # Filtrera investeringar/utrangeringar
    invest = capbase_df[capbase_df['metod'] == 'future_invest'].copy()
    
    if invest.empty:
        return pd.DataFrame()
    
    # Skapa KENT-format
    kent_invest = pd.DataFrame()
    
    # Investering / Utrangering
    kent_invest['Investering / Utrangering'] = invest['invest'].map({
        1: 'Investering',
        -1: 'Utrangering'
    })
    
    # Halvår
    def format_halfyear(time_val):
        year, h = time_to_year_halfyear(time_val)
        if year is None:
            return ''
        return f"{year} H{h}"
    
    kent_invest['Halvår'] = invest['time_invest'].apply(format_halfyear)
    
    # Anl.kategori och Typ av anläggning
    kent_invest['Anl.kategori'] = invest['cat']
    kent_invest['Typ av anläggning'] = invest['subcat']
    
    # Antal
    kent_invest['Antal'] = invest['count_comp']
    
    # Ursprungligen tagen i bruk (om relevant)
    kent_invest['Ursprungligen tagen i bruk'] = invest['time_from'].apply(
        lambda x: time_to_year(x) if pd.notna(x) else ''
    )
    
    # Totalt i kronor (ta absolutvärdet)
    kent_invest['Totalt i kronor'] = invest['nuav_2022'].abs()
    
    # Anmärkning
    kent_invest['Anmärkning'] = ''
    
    return kent_invest


def format_kent_excel(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    header_row: int = 2
):
    """Formaterar ett KENT-ark med korrekt stil."""
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    # Rensa arket
    ws.delete_rows(1, ws.max_row)
    
    # Lägg till titel
    ws.cell(row=1, column=1, value=f"KENT - {sheet_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Lägg till data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=header_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formatera header
            if r_idx == header_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Justera kolumnbredd
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_kent_from_capbase(
    capbase_path: str,
    uppslagsvarden_path: str,
    output_path: str,
    network_id: Optional[int] = None
) -> Dict[str, any]:
    """
    Huvudfunktion: Skapar KENT Excel-fil från capbase_a.
    
    Args:
        capbase_path: Path till capbase_a.parquet
        uppslagsvarden_path: Path till uppslagsvärden Excel
        output_path: Path för output KENT-fil
        network_id: Filtera på specifikt nätverks-ID (None = alla)
    
    Returns:
        Dictionary med statistik och resultat
    """
    
    # Läs capbase_a
    capbase_df = pd.read_parquet(capbase_path)
    
    # Filtrera på network_id om specificerat
    if network_id is not None:
        capbase_df = capbase_df[capbase_df['id_network'] == network_id].copy()
    
    # Läs uppslagsvärden
    uppslagsvarden_df = load_uppslagsvarden(uppslagsvarden_path)
    
    # Skapa ark
    normvarde_df = create_normvarde_sheet(capbase_df, uppslagsvarden_df)
    ovriga_df = create_ovriga_metoder_sheet(capbase_df)
    invest_df = create_investeringar_sheet(capbase_df)
    
    # Skapa Excel-fil från scratch
    wb = Workbook()
    # Ta bort default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Lägg till nya ark
    if not normvarde_df.empty:
        format_kent_excel(wb, 'Normvärde', normvarde_df)
    
    if not ovriga_df.empty:
        format_kent_excel(wb, 'Övriga värderingsmetoder', ovriga_df)
    
    if not invest_df.empty:
        format_kent_excel(wb, 'Investeringar_Utrangeringar', invest_df)
    
    # Lägg till Uppslagsvärden från original-filen
    wb_orig = load_workbook(uppslagsvarden_path)
    if 'Uppslagsvärden' in wb_orig.sheetnames:
        ws_orig = wb_orig['Uppslagsvärden']
        ws_new = wb.create_sheet('Uppslagsvärden')
        
        # Lägg till en tom rad först (så header=1 fungerar i capbase_prep.py)
        ws_new.cell(row=1, column=1, value='')
        
        # Kopiera alla celler från rad 2 och framåt
        for row_idx, row in enumerate(ws_orig.iter_rows(), start=2):
            for cell in row:
                new_cell = ws_new.cell(row=row_idx, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    try:
                        new_cell.font = cell.font.copy() if hasattr(cell.font, 'copy') else cell.font
                        new_cell.border = cell.border.copy() if hasattr(cell.border, 'copy') else cell.border
                        new_cell.fill = cell.fill.copy() if hasattr(cell.fill, 'copy') else cell.fill
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = cell.alignment.copy() if hasattr(cell.alignment, 'copy') else cell.alignment
                    except:
                        pass
        
        # Kopiera kolumnbredder
        for col_letter, col_dim in ws_orig.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = col_dim.width
    
    # Ordna om arken i rätt ordning
    sheet_order = ['Normvärde', 'Övriga värderingsmetoder', 
                   'Investeringar_Utrangeringar', 'Uppslagsvärden']
    
    for idx, sheet_name in enumerate(sheet_order):
        if sheet_name in wb.sheetnames:
            # Flytta arket till rätt position
            current_idx = wb.sheetnames.index(sheet_name)
            if current_idx != idx:
                wb.move_sheet(sheet_name, offset=(idx - current_idx))
    
    # Spara
    wb.save(output_path)
    
    # Statistik
    stats = {
        'total_components': len(capbase_df),
        'normvarde_count': len(normvarde_df),
        'ovriga_count': len(ovriga_df),
        'investeringar_count': len(invest_df),
        'network_id': network_id or 'all',
        'output_file': output_path
    }
    
    return stats


def validate_round_trip(
    kent_file: str,
    original_capbase: str,
    capbase_prep_function
) -> Dict[str, any]:
    """
    Validerar att KENT-fil → capbase_a ger samma resultat.
    
    Args:
        kent_file: Skapad KENT-fil
        original_capbase: Original capbase_a.parquet
        capbase_prep_function: Funktion från capbase_prep.py
    
    Returns:
        Valideringsrapport
    """
    
    # Läs original
    original_df = pd.read_parquet(original_capbase)
    
    # Generera från KENT
    reconstructed_df = capbase_prep_function(kent_file)
    
    # Jämför
    report = {
        'row_count_match': len(original_df) == len(reconstructed_df),
        'original_rows': len(original_df),
        'reconstructed_rows': len(reconstructed_df),
        'column_differences': [],
        'value_differences': []
    }
    
    # Kolumnjämförelse
    common_cols = set(original_df.columns) & set(reconstructed_df.columns)
    
    for col in common_cols:
        if col in ['id_component']:  # Skip auto-generated IDs
            continue
        
        # Numerisk jämförelse
        if pd.api.types.is_numeric_dtype(original_df[col]):
            diff = (original_df[col] - reconstructed_df[col]).abs()
            max_diff = diff.max()
            if max_diff > 0.01:  # Tolerance för float-precision
                report['value_differences'].append({
                    'column': col,
                    'max_difference': float(max_diff),
                    'mean_difference': float(diff.mean())
                })
    
    return report


if __name__ == '__main__':
    """Exempel på användning."""
    
    # Paths
    base_dir = Path(__file__).parent.parent
    capbase_path = base_dir / 'data' / 'capbase_a.parquet'
    uppslagsvarden_path = base_dir / 'data' / 'KENT_Inrapporteringsmall_index_och_uppslagsvarden.xlsx'
    output_path = base_dir / 'data' / 'KENT_reconstructed.xlsx'
    
    # Skapa KENT-fil
    print("Skapar KENT-fil från capbase_a...")
    stats = create_kent_from_capbase(
        str(capbase_path),
        str(uppslagsvarden_path),
        str(output_path),
        network_id=886  # Filtrera på nätverks-ID 886
    )
    
    print("\n=== RESULTAT ===")
    print(f"Total komponenter: {stats['total_components']}")
    print(f"Normvärde: {stats['normvarde_count']}")
    print(f"Övriga metoder: {stats['ovriga_count']}")
    print(f"Investeringar: {stats['investeringar_count']}")
    print(f"\nKENT-fil sparad: {stats['output_file']}")