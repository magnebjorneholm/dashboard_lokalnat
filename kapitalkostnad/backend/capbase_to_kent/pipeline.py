"""
capbase_to_kent_backend.py - Backend för konvertering av capbase_a till KENT

Läser capbase_a.parquet och skapar KENT_reconstructed.xlsx i samma mapp.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def time_to_year_halfyear(time_value):
    """Konverterar time till år och halvår."""
    if pd.isna(time_value):
        return None, None
    year = int((time_value - 1) // 2 + 1910)
    halfyear = int((time_value - 1) % 2 + 1)
    return year, halfyear


def time_to_year(time_value):
    """Konverterar time till år."""
    year, _ = time_to_year_halfyear(time_value)
    return year


def create_normvarde_sheet(capbase_df, uppslagsvarden_df):
    """Skapar Normvärde-ark."""
    normvarde = capbase_df[capbase_df['metod'] == 'normvärde'].copy()
    if normvarde.empty:
        return pd.DataFrame()
    
    uppslagsvarden_mapping = uppslagsvarden_df.set_index('Kod')[
        ['Anläggningskategori', 'Typ av anläggning', 'Teknisk specifikation', 
         'Spänning kV', 'Normvärde 2022 (SEK)']
    ].to_dict('index')
    
    kent_normvarde = pd.DataFrame()
    kent_normvarde['Anl.-kategori'] = normvarde['id_comptype'].apply(
        lambda kod: uppslagsvarden_mapping.get(kod, {}).get('Anläggningskategori', '')
    )
    kent_normvarde['Kod'] = normvarde['id_comptype']
    kent_normvarde['Typ av anläggning'] = normvarde['id_comptype'].apply(
        lambda kod: uppslagsvarden_mapping.get(kod, {}).get('Typ av anläggning', '')
    )
    kent_normvarde['Teknisk specifikation'] = normvarde['id_comptype'].apply(
        lambda kod: uppslagsvarden_mapping.get(kod, {}).get('Teknisk specifikation', '')
    )
    kent_normvarde['Spänning'] = normvarde['id_comptype'].apply(
        lambda kod: uppslagsvarden_mapping.get(kod, {}).get('Spänning kV', '')
    )
    kent_normvarde['Antal'] = normvarde['count_comp']
    kent_normvarde['Rådighet'] = normvarde['owned'].map({1: 'Ägd', 0: 'Hyrd/Leasad'})
    kent_normvarde['Ursprungligen tagen i bruk'] = normvarde['time_from'].apply(time_to_year)
    kent_normvarde['NUAV (kr)'] = normvarde['id_comptype'].apply(
        lambda kod: uppslagsvarden_mapping.get(kod, {}).get('Normvärde 2022 (SEK)', 0)
    ) * normvarde['count_comp'].values
    
    mask_missing = normvarde['time_from_missing'] == 1
    if mask_missing.any():
        kent_normvarde.loc[mask_missing, 'År saknas (Ja eller blank)'] = 'Ja'
        kent_normvarde.loc[mask_missing, 'Ursprungligen tagen i bruk'] = ''
    
    kent_normvarde['Anmärkning'] = ''
    return kent_normvarde


def create_ovriga_metoder_sheet(capbase_df):
    """Skapar Övriga värderingsmetoder-ark."""
    ovriga = capbase_df[
        capbase_df['metod'].isin(['anskaffningsvärde', 'bokförtvärde', 'annatskäligtvärde'])
    ].copy()
    if ovriga.empty:
        return pd.DataFrame()
    
    kent_ovriga = pd.DataFrame()
    kent_ovriga['Ansk'] = ovriga['metod'].apply(lambda x: 'x' if x == 'anskaffningsvärde' else '')
    kent_ovriga['Bokf'] = ovriga['metod'].apply(lambda x: 'x' if x == 'bokförtvärde' else '')
    kent_ovriga['Annat'] = ovriga['metod'].apply(lambda x: 'x' if x == 'annatskäligtvärde' else '')
    kent_ovriga['Anl.kategori'] = ovriga['cat']
    kent_ovriga['Typ av anläggning'] = ovriga['subcat']
    kent_ovriga['Antal'] = ovriga['count_comp']
    kent_ovriga['Ursprungligen tagen i bruk'] = ovriga['time_from'].apply(time_to_year)
    kent_ovriga['Rådighet'] = ovriga['owned'].map({1: 'Ägd', 0: 'Hyrd/Leasad'})
    kent_ovriga['NUAV 2022 (kr)'] = ovriga['nuav_2022']
    
    mask_missing = ovriga['time_from_missing'] == 1
    if mask_missing.any():
        kent_ovriga.loc[mask_missing, 'År saknas (Ja eller blank)'] = 'Ja'
        kent_ovriga.loc[mask_missing, 'Ursprungligen tagen i bruk'] = ''
    
    kent_ovriga['Anmärkning'] = ''
    return kent_ovriga


def create_investeringar_sheet(capbase_df):
    """Skapar Investeringar_Utrangeringar-ark."""
    invest = capbase_df[capbase_df['metod'] == 'future_invest'].copy()
    if invest.empty:
        return pd.DataFrame()
    
    kent_invest = pd.DataFrame()
    kent_invest['Investering / Utrangering'] = invest['invest'].map({
        1: 'Investering',
        -1: 'Utrangering'
    })
    
    def format_halfyear(time_val):
        year, h = time_to_year_halfyear(time_val)
        if year is None:
            return ''
        return f"{year} H{h}"
    
    kent_invest['Halvår'] = invest['time_invest'].apply(format_halfyear)
    kent_invest['Anl.kategori'] = invest['cat']
    kent_invest['Typ av anläggning'] = invest['subcat']
    kent_invest['Antal'] = invest['count_comp']
    kent_invest['Ursprungligen tagen i bruk'] = invest['time_from'].apply(
        lambda x: time_to_year(x) if pd.notna(x) else ''
    )
    kent_invest['Totalt i kronor'] = invest['nuav_2022'].abs()
    kent_invest['Anmärkning'] = ''
    return kent_invest


def format_kent_excel(wb, sheet_name, df, header_row=2):
    """Formaterar KENT-ark."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    ws.delete_rows(1, ws.max_row)
    
    ws.cell(row=1, column=1, value=f"KENT - {sheet_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=header_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == header_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def load_uppslagsvarden(uppslagsvarden_path):
    """Läser uppslagsvärden."""
    df = pd.read_excel(uppslagsvarden_path, sheet_name='Uppslagsvärden', header=0)
    df = df[['Anläggningskategori', 'Kod', 'Typ av anläggning', 
             'Teknisk specifikation', 'Spänning kV', 'Normvärde 2022 (SEK)']].copy()
    df = df[df['Kod'].notna()].copy()
    return df


def capbase_to_kent(capbase_path, network_id=None):
    """
    Huvudfunktion: Konverterar capbase_a.parquet till KENT_reconstructed.xlsx
    
    Args:
        capbase_path: Path till capbase_a.parquet
        network_id: Valfritt - filtrera på specifikt nätverks-ID
    """
    
    capbase_path = Path(capbase_path)
    output_dir = capbase_path.parent
    
    capbase_df = pd.read_parquet(capbase_path)
    if network_id is not None:
        capbase_df = capbase_df[capbase_df['id_network'] == network_id].copy()
    
    uppslagsvarden_path = Path(__file__).parent.parent.parent / 'data' / 'KENT_Inrapporteringsmall_index_och_uppslagsvarden.xlsx'
    uppslagsvarden_df = load_uppslagsvarden(uppslagsvarden_path)
    
    normvarde_df = create_normvarde_sheet(capbase_df, uppslagsvarden_df)
    ovriga_df = create_ovriga_metoder_sheet(capbase_df)
    invest_df = create_investeringar_sheet(capbase_df)
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    if not normvarde_df.empty:
        format_kent_excel(wb, 'Normvärde', normvarde_df)
    
    if not ovriga_df.empty:
        format_kent_excel(wb, 'Övriga värderingsmetoder', ovriga_df)
    
    if not invest_df.empty:
        format_kent_excel(wb, 'Investeringar_Utrangeringar', invest_df)
    
    wb_orig = load_workbook(uppslagsvarden_path)
    if 'Uppslagsvärden' in wb_orig.sheetnames:
        ws_orig = wb_orig['Uppslagsvärden']
        ws_new = wb.create_sheet('Uppslagsvärden')
        ws_new.cell(row=1, column=1, value='')
        
        for row_idx, row in enumerate(ws_orig.iter_rows(), start=2):
            for cell in row:
                new_cell = ws_new.cell(row=row_idx, column=cell.column)
                new_cell.value = cell.value
        
        for col_letter, col_dim in ws_orig.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = col_dim.width
    
    sheet_order = ['Normvärde', 'Övriga värderingsmetoder', 
                   'Investeringar_Utrangeringar', 'Uppslagsvärden']
    for idx, sheet_name in enumerate(sheet_order):
        if sheet_name in wb.sheetnames:
            current_idx = wb.sheetnames.index(sheet_name)
            if current_idx != idx:
                wb.move_sheet(sheet_name, offset=(idx - current_idx))
    
    output_path = output_dir / 'KENT_reconstructed.xlsx'
    wb.save(output_path)
    
    return str(output_path)


if __name__ == '__main__':
    
    capbase_path = Path(__file__).parent.parent.parent / 'data' / 'capbase_a.parquet'
    
    print("Konverterar capbase_a.parquet → KENT_reconstructed.xlsx...")
    output_path = capbase_to_kent(capbase_path, network_id=886)
    print(f"✓ Klar! Sparad: {output_path}")